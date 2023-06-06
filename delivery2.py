import pandas as pd
import openpyxl as pyxl
import numpy as np
import time

#file_path = "C:/Users/fun-f/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務フォルダ/③商品部/入荷管理/納品スケジュール/2022/納品スケジュール【2022】.xlsx"
#file_path = "C:/Users/fun-f/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務フォルダ/③商品部/入荷管理/納品スケジュール/2023/納品スケジュール【2023】.xlsx"
file_path = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務フォルダ/③商品部/入荷管理/納品スケジュール/2023/納品スケジュール【2023】.xlsx"
#file_path = 'C:/Users/fun-f/Downloads/納品スケジュール【2023】.xlsx'
tenpo_list_1 = {

"柏":"",
"千葉":"",
"伊勢崎": "",
# "長町":"",
# "TOKYO-BAY":"",
"富士見":"",
"越谷":"",
"海老名":"",
"むさし村山":"",
"湘南平塚":"",
"名取":"",
"大高":"",
"愛知東郷":"",
"太田":"",
"水戸内原":"",
"EXPOCITY":"",
"川崎":"",
"新三郷":"",
"幕張新都心":"",
"各務原":"",
"堺":"",

}

#r_file = pd.read_excel(file_path)

r_file = pyxl.load_workbook(file_path)

ws_names = r_file.sheetnames


select_list = []

key_nos = 0
for ws_name in ws_names:
  
  if ws_name == "納品スケジュール　原紙":
    print("No_count")
  else:
  #select_list.append([key_nos,ws_name])
    select_list.append(ws_name)
   
  key_nos += 1
  
print("シート名一覧\n" + str(select_list))

print("選択No.を入力して下さい！")
select_1 = int(input())

print("『 " + str(select_list[select_1]) + " 』のシートを選択します！")

print("開始しますか？\n\nYES ⇒ 0\nNo ⇒ 1")

select_2 = int(input())

if select_2 == 0 :
  print("開始………")

  ws = r_file[str(select_list[select_1])]
  
  import datetime
  import requests

  import selenium
  from selenium import webdriver
  import time
  from datetime import timedelta
  from selenium.webdriver.chrome.options import Options
  from selenium.webdriver.common.by import By
  from selenium.webdriver.support.ui import Select
  chrome_options = Options()
  chrome_options.add_experimental_option("detach", True)

  #ーーーーーーー販売NETスクレイピングーーーーーーーーーーー

  week = ['月','火','水','木','金','土','日']
  w_day = '{:%Y%m%d}'.format(datetime.datetime.today())

  year = w_day[0:4]
  month = w_day[4:6]
  #day = int(w_day[6:8]) + 1#変更前
  day = w_day[6:8] #+ 1
  #week_day_type = datetime.date(int(year), int(month), int(day)).isocalendar()[2] - 1
  w_day_df = datetime.datetime.strptime(str(year) + '-' + str(month) + '-' + str(day), '%Y-%m-%d')
  
  day = w_day_df + timedelta(days = 1)
  year = str(day.year).zfill(4)
  month = str(day.month).zfill(2)
  day = str(day.day).zfill(2)
  target_day = str(year) + str(month)+ str(day).zfill(2)
  
  

  #week_no = datetime.date(int(year), int(month), int(day)).isocalendar()[1] + 1
  week_no = datetime.date(int(year), int(2), int(2)).isocalendar()[1] +1
  print(week_no)  # 53
  week_day_type = datetime.date(int(year), int(month), int(day)).isocalendar()[2] - 1
  #week_day_type = datetime.date(int(year), int(2), int(2)).isocalendar()[2] 
  
  print(week_day_type)
  print(week[week_day_type])

  url = 'http://tri.hanbai-net.com/system/Login.aspx'
  #driver = webdriver.Chrome('C:/Users/fun-f/Downloads/chromedriver.exe')#旧
  #driver = webdriver.Chrome('C:/Users/fun-f/Desktop/myfile/chromedriver.exe')
  from webdriver_manager.chrome import ChromeDriverManager
 
  #driver = webdriver.Chrome(ChromeDriverManager().install())
  #driver = webdriver.Chrome("C:/Users/古内翔平/chromedriver.exe")#2021 0724
  driver = webdriver.Chrome(options=chrome_options)#2021 0724

  #driver = webdriver.Chrome("C:/Users/fun-f/chromedriver.exe")#2021 0724

  driver.get(url)

  #id_1 = 'tenpo'
  #id_2 = 'tenpo'
  
  id_1 = 'trinityadmin'
  id_2 = 'AdminTrinity'
  
  loginid_1 = driver.find_element(By.ID, "ContentPlaceHolder1_txtUserCode")
  loginid_2 = driver.find_element(By.ID, "ContentPlaceHolder1_txtPassword")

  loginid_1.send_keys(id_1)#ユーザーIDを入力
  loginid_2.send_keys(id_2)#パスワードを入力



  driver.find_element(By.ID,"ContentPlaceHolder1_btnLogin").click() 
  #ログインボタンをクリック

  driver.get('http://tri.hanbai-net.com/system/00000000.aspx')
  driver.get('http://tri.hanbai-net.com/system/21024101.aspx?id=010199')

  driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02").send_keys(target_day)#開始日を指定 ⇒ w_day
  driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02_2").send_keys(target_day)#末日を指定 ⇒ w_day
  driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton6").click()

  tenpo_list = {

  '柏':"01001008",
  '千葉':"01001009",
  '伊勢崎':"01001028",
  # '長町':'//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[5]',
  # 'TOKYO-BAY':'//*[@id="ContentPlaceHolder1_DropDownListCond04"]/option[6]',
  '富士見':"01001034",#7
  '越谷':"01001036",
  '海老名':"01001038",
  'むさし村山':"01001039",
  '湘南平塚':"01001040",
  '名取':"01001041",
  '大高':"01001042",
  '愛知東郷':"01001043",
  '太田':"01001044",
  '水戸内原':"01001045",
  'EXPOCITY':"01001046",
  '川崎':"01001047",
  '新三郷':"01001048",
  '幕張新都心':"01001049",
  '各務原':"01001050",
  '堺':"01001051"
  

  }
  
  col_list = {
    
    
    0:["C","E","G","I","K","M","O","Q","S","U"],#13
    1:["C","E","G","I","K","M","O","Q","S","U"],#50
    2:["C","E","G","I","K","M","O","Q","S","U"],#87
    3:["AC","AE","AG","AI","AK","AM","AO","AQ","AS","AU"],#13
    
    # 4:["AC","AE","AG","AI","AK","AM","AO","AQ","AS","AU"],#50
    
    4:["AC","AE","AG","AI","AK","AM","AO","AQ","AS","AU"],#50
    
    5:["AC","AE","AG","AI","AK","AM","AO","AQ","AS","AU"],#87

  }
  

  file_name = '納品スケジュール【' + str(year) + '】.xlsx'


  # path = r'C:/Users/fun-f/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務フォルダ/③商品部/入荷管理/納品スケジュール/2022/'
  # path = r'C:/Users/fun-f/Downloads/納品スケジュール【2023】.xlsx'
  #path = r"C:/Users/fun-f/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務フォルダ/③商品部/入荷管理/納品スケジュール/2023/納品スケジュール【2023】.xlsx"
  path = r"C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務フォルダ/③商品部/入荷管理/納品スケジュール/2023/納品スケジュール【2023】.xlsx"
  

  #sheet_name = '納品スケジュール' + str(year) + '.'+ str(int(month)) + '.' + str(week_no) + 'W'
  sheet_name = select_list[select_1]
  print(sheet_name)
  #wb1 = pyxl.load_workbook(path + file_name,data_only=True)
  wb1 = pyxl.load_workbook(path,data_only=True)

  act_ws_1 = wb1[sheet_name]

  p_avg = act_ws_1.cell(row=3,column=7).value
  p_list = []
  
  counter = 19 #店舗数-1の数を入力する
  
  if week_day_type == 0:
    row_no = 13
    mfr_list = []

    #月
    
    #------------------------------------------------------------------------------------------------------------------------------------
    #追加
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton10").click()#店間移動にチェック
    
    time.sleep(2)
    
    driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()
    
    time.sleep(5)
    
    mbs = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text
    mbs_2 = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
    
    print(mbs)
    
    if str(mbs) == '-':
      
      mbs_message = "◆ 【店間 - 全店合計数】\n" + "なし"
      
    elif int(''.join(filter(str.isdigit, mbs))) > 0:
      mbs_message = "◆ 【店間 - 全店合計数】\n" + str(mbs) + "点 / " + str(mbs_2) 
      
    driver.get('http://tri.hanbai-net.com/system/21024101.aspx?id=010199')

    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02").send_keys(target_day)#開始日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02_2").send_keys(target_day)#末日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton6").click()  
      
    for n in range(0,counter):
      
      shop_name1 = act_ws_1.cell(row=15 + n,column=2).value
      quantity1 = act_ws_1.cell(row=15 + n,column=33).value#納品点数
      p1 = act_ws_1.cell(row=15 + n,column=34).value#パッキン数
      #---------------------------------------------------------------------------
      shop_input = driver.find_element(By.ID,"ContentPlaceHolder1_DropDownListCond04")#.click()
      shop_input_select = Select(shop_input)
      shop_input_select.select_by_value(tenpo_list[shop_name1])

      driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()

      time.sleep(3)

      p_ = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
      p_f = ''.join(filter(str.isdigit, p_))
      

      print(p_f)

      q_ = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text

      print(q_)
      
      quantity1 = int(quantity1) + int(q_) #納品点数
      p1 = float(p1) + float(p_f) #パッキン数
      
      #---------------------------------------------------------------------------
      
      if p1 == None:
        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(0) +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(0)]})
        p_list.append(row1)
    
      else :

        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(f"{p1: .1f}") +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(f"{p1: .1f}")]})
        p_list.append(row1)
      
    #納品メーカー名を取得  

    for c in range(0,10):
      mfr = act_ws_1[str(col_list[int(week_day_type)][int(c)]) + str(row_no)].value
    
      if mfr == None:
        print("No Maker")
      
      else:
        mfr_list.append(mfr)
        
        
    mfr_count = len(mfr_list)
    print("メーカー数"+ str(mfr_count))
    print(mfr_list)
    
    
    try :
      mfr_1 = mfr_list[0]
    except IndexError:
      print("NoData")
      mfr_1 = ""
      
    try:  
      mfr_2 = mfr_list[1]
      
    except IndexError:
      print("NoData")
      mfr_2 = ""
      
    try:
        
      mfr_3 = mfr_list[2]
    except IndexError:
      print("NoData")  
      mfr_3 = ""
      
    try:  
      mfr_4 = mfr_list[3]
      
    except IndexError:
      print("NoDAta")
      
      mfr_4 = ""
      
    try:  
      
      mfr_5 = mfr_list[4]
      
    except IndexError:  
      
      print("NoData")
      
      mfr_5 = ""
      
    try:
        
      mfr_6 = mfr_list[5]
    except IndexError:
      print("NoData")  
      mfr_6 = ""
      
    try:
        
      mfr_7 = mfr_list[6]
      
    except IndexError:  
      print("NoData")
      mfr_7 = ""
      
    try:  
      mfr_8 = mfr_list[7]
      
    except IndexError:
      
      print("NoData")
      mfr_8 = ""
      
    try:  
      mfr_9 = mfr_list[8]
      
    except IndexError:
      print("NoData")  
      mfr_9 = ""
      
    try:  
      mfr_10 = mfr_list[9] 
    except IndexError:
      print("NoData")
      mfr_10 = ""
          
        
  elif week_day_type == 1 :
    
    row_no = 50
    mfr_list = []

    #火
    
    #------------------------------------------------------------------------------------------------------------------------------------
    #追加
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton10").click()#店間移動にチェック
    
    time.sleep(2)
    
    driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()
    
    time.sleep(5)
    
    mbs = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text
    mbs_2 = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
    
    print(mbs)
    
    if str(mbs) == '-':
      
      mbs_message = "◆ 【店間 - 全店合計数】\n" + "なし"
      
    elif int(''.join(filter(str.isdigit, mbs))) > 0:
      mbs_message = "◆ 【店間 - 全店合計数】\n" + str(mbs) + "点 / " + str(mbs_2) 
      
    driver.get('http://tri.hanbai-net.com/system/21024101.aspx?id=010199')

    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond02"]').send_keys(target_day)#開始日を指定 ⇒ w_day
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_txtCond02_2"]').send_keys(target_day)#末日を指定 ⇒ w_day
    driver.find_element_by_xpath('//*[@id="ContentPlaceHolder1_RadioButton6"]').click()  
      
    for n in range(0,counter):
      shop_name1 = act_ws_1.cell(row=52 + n,column=2).value
      quantity1 = act_ws_1.cell(row=52 + n,column=23).value#納品点数
      p1 = act_ws_1.cell(row=52 + n,column=24).value#パッキン数
      #---------------------------------------------------------------------------
      
      
      #------------------------------------------------------------------------------------------------------------------------------------
        
      shop_input = driver.find_element(By.ID,"ContentPlaceHolder1_DropDownListCond04")#.click()
      shop_input_select = Select(shop_input)
      shop_input_select.select_by_value(tenpo_list[shop_name1])

      driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()

      time.sleep(2)

      p_ = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
      p_f = ''.join(filter(str.isdigit, p_))

      print(p_f)

      q_ = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text

      print(q_)
      
      quantity1 = int(quantity1) + int(q_) #納品点数
      
      try:
        p1 = float(p1) + float(p_f) #パッキン数
        
      except TypeError:  
        p1 = int(0) + int(0) #パッキン数
      
      #---------------------------------------------------------------------------
      
      if p1 == None:
        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(0) +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(0)]})
        p_list.append(row1)
    
      else :

        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(f"{p1: .1f}") +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(f"{p1: .1f}")]})
        p_list.append(row1)
    
    #納品メーカー名を取得  

    for c in range(0,10):
      mfr = act_ws_1[str(col_list[int(week_day_type)][int(c)]) + str(row_no)].value
    
      if mfr == None:
        print("No Maker")
      
      else:
        mfr_list.append(mfr)
        
        
    mfr_count = len(mfr_list)
    print("メーカー数"+ str(mfr_count))
    print(mfr_list)
    
    
    try :
      mfr_1 = mfr_list[0]
    except IndexError:
      print("NoData")
      mfr_1 = ""
      
    try:  
      mfr_2 = mfr_list[1]
      
    except IndexError:
      print("NoData")
      mfr_2 = ""
      
    try:
        
      mfr_3 = mfr_list[2]
    except IndexError:
      print("NoData")  
      mfr_3 = ""
      
    try:  
      mfr_4 = mfr_list[3]
      
    except IndexError:
      print("NoDAta")
      
      mfr_4 = ""
      
    try:  
      
      mfr_5 = mfr_list[4]
      
    except IndexError:  
      
      print("NoData")
      
      mfr_5 = ""
      
    try:
        
      mfr_6 = mfr_list[5]
    except IndexError:
      print("NoData")  
      mfr_6 = ""
      
    try:
        
      mfr_7 = mfr_list[6]
      
    except IndexError:  
      print("NoData")
      mfr_7 = ""
      
    try:  
      mfr_8 = mfr_list[7]
      
    except IndexError:
      
      print("NoData")
      mfr_8 = ""
      
    try:  
      mfr_9 = mfr_list[8]
      
    except IndexError:
      print("NoData")  
      mfr_9 = ""
      
    try:  
      mfr_10 = mfr_list[9] 
    except IndexError:
      print("NoData")
      mfr_10 = ""
          

  elif week_day_type == 2 :
    
    row_no = 87
    mfr_list = []
    #水
    
    #------------------------------------------------------------------------------------------------------------------------------------
    #追加ContentPlaceHolder1_RadioButton10
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton10").click()#店間移動にチェック
    
    time.sleep(2)
    
    driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()
    
    time.sleep(5)
    
    mbs = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text
    mbs_2 = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
    
    print(mbs)
    
    if str(mbs) == '-':
      
      mbs_message = "◆ 【店間 - 全店合計数】\n" + "なし"
      
    elif int(''.join(filter(str.isdigit, mbs))) > 0:
      mbs_message = "◆ 【店間 - 全店合計数】\n" + str(mbs) + "点 / " + str(mbs_2) 
      
    driver.get('http://tri.hanbai-net.com/system/21024101.aspx?id=010199')

    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02").send_keys(target_day)#開始日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02_2").send_keys(target_day)#末日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton6").click()  
      
    for n in range(0,counter):
      shop_name1 = act_ws_1.cell(row=89 + n,column=2).value
      quantity1 = act_ws_1.cell(row=89 + n,column=23).value#納品点数
      p1 = act_ws_1.cell(row=89 + n,column=24).value#パッキン数
      #---------------------------------------------------------------------------
      
      
      #------------------------------------------------------------------------------------------------------------------------------------
        
      shop_input = driver.find_element(By.ID,"ContentPlaceHolder1_DropDownListCond04")#.click()
      shop_input_select = Select(shop_input)
      shop_input_select.select_by_value(tenpo_list[shop_name1])

      driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()

      time.sleep(2)

      p_ = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
      p_f = ''.join(filter(str.isdigit, p_))

      print(p_f)

      q_ = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text

      print(q_)
      
      quantity1 = int(quantity1) + int(q_) #納品点数
      #p1 = float(p1) + float(p_f) #パッキン数
      
      try:
        p1 = float(p1) + float(p_f) #パッキン数
        
      except TypeError:  
        p1 = int(0) + int(0) #パッキン数
      
      #---------------------------------------------------------------------------
      
    
      if p1 == None:
        
        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(0) +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(0)]})
        p_list.append(row1)
    
        
      else :

        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(f"{p1: .1f}") +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(f"{p1: .1f}")]})
        p_list.append(row1)
        
    #納品メーカー名を取得  

    for c in range(0,10):
      mfr = act_ws_1[str(col_list[int(week_day_type)][int(c)]) + str(row_no)].value
    
      if mfr == None:
        print("No Maker")
      
      else:
        mfr_list.append(mfr)
        
        
    mfr_count = len(mfr_list)
    print("メーカー数"+ str(mfr_count))
    print(mfr_list)
    
    
    try :
      mfr_1 = mfr_list[0]
    except IndexError:
      print("NoData")
      mfr_1 = ""
      
    try:  
      mfr_2 = mfr_list[1]
      
    except IndexError:
      print("NoData")
      mfr_2 = ""
      
    try:
        
      mfr_3 = mfr_list[2]
    except IndexError:
      print("NoData")  
      mfr_3 = ""
      
    try:  
      mfr_4 = mfr_list[3]
      
    except IndexError:
      print("NoDAta")
      
      mfr_4 = ""
      
    try:  
      
      mfr_5 = mfr_list[4]
      
    except IndexError:  
      
      print("NoData")
      
      mfr_5 = ""
      
    try:
        
      mfr_6 = mfr_list[5]
    except IndexError:
      print("NoData")  
      mfr_6 = ""
      
    try:
        
      mfr_7 = mfr_list[6]
      
    except IndexError:  
      print("NoData")
      mfr_7 = ""
      
    try:  
      mfr_8 = mfr_list[7]
      
    except IndexError:
      
      print("NoData")
      mfr_8 = ""
      
    try:  
      mfr_9 = mfr_list[8]
      
    except IndexError:
      print("NoData")  
      mfr_9 = ""
      
    try:  
      mfr_10 = mfr_list[9] 
    except IndexError:
      print("NoData")
      mfr_10 = ""
      
    try:  
      mfr_11 = mfr_list[10] 
    except IndexError:
      print("NoData")
      mfr_11 = ""
      
    try:  
      mfr_12 = mfr_list[11] 
    except IndexError:
      print("NoData")
      mfr_12 = ""   
      
    try:  
      mfr_13 = mfr_list[12] 
    except IndexError:
      print("NoData")
      mfr_13 = ""
      
    try:  
      mfr_14 = mfr_list[13] 
    except IndexError:
      print("NoData")
      mfr_14 = ""     
      
    try:  
      mfr_15 = mfr_list[14] 
    except IndexError:
      print("NoData")
      mfr_15 = ""  
            
        
  elif week_day_type == 3 :  
    
    row_no = 13
    mfr_list = []  
  
    #木
    
    #------------------------------------------------------------------------------------------------------------------------------------
    #追加
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton10").click()#店間移動にチェック
    
    time.sleep(2)
    
    driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()
    
    time.sleep(5)
    
    mbs = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text
    mbs_2 = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
 
    
    print(mbs)
    
    if str(mbs) == '-':
      
      mbs_message = "◆ 【店間 - 全店合計数】\n" + "なし"
      
    elif int(''.join(filter(str.isdigit, mbs))) > 0:
      mbs_message = "◆ 【店間 - 全店合計数】\n" + str(mbs) + "点 / " + str(mbs_2) 
      
    driver.get('http://tri.hanbai-net.com/system/21024101.aspx?id=010199')

    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02").send_keys(target_day)#開始日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02_2").send_keys(target_day)#末日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton6").click()  
      
      
    for n in range(0,counter):
      shop_name1 = act_ws_1.cell(row=15 + n,column=28).value
      quantity1 = act_ws_1.cell(row=15 + n,column=49).value#納品点数
      p1 = act_ws_1.cell(row=15 + n,column=50).value#パッキン数
      
      
      #------------------------------------------------------------------------------------------------------------------------------------
        
      #driver.find_element(By.ID,"ContentPlaceHolder1_UpdatePanel2").click()#店間移動にチェック  
      #------------------------------------------------------------------------------------------------------------------------------------
      
      
      shop_input = driver.find_element(By.ID,"ContentPlaceHolder1_DropDownListCond04")#.click()
      shop_input_select = Select(shop_input)
      shop_input_select.select_by_value(tenpo_list[shop_name1])
  
      #driver.find_element(tenpo_list[shop_name1]).click()

      driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()

      time.sleep(2)

      p_ = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
      p_f = ''.join(filter(str.isdigit, p_))

      print(p_f)

      q_ = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text

      print(q_)
      
      quantity1 = int(quantity1) + int(q_) #納品点数
      #p1 = float(p1) + float(p_f) #パッキン数
      
      try:
        p1 = float(p1) + float(p_f) #パッキン数
        
      except TypeError:  
        p1 = int(0) + int(0) #パッキン数
      
      #---------------------------------------------------------------------------
    
      if p1 == None:
        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(0) +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(0)]})
        p_list.append(row1)
        
      else :

        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(f"{p1: .1f}") +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(f"{p1: .1f}")]})
        p_list.append(row1)
        
        
    #納品メーカー名を取得  

    for c in range(0,10):
      mfr = act_ws_1[str(col_list[int(week_day_type)][int(c)]) + str(row_no)].value
    
      if mfr == None:
        print("No Maker")
      
      else:
        mfr_list.append(mfr)
        
        
    mfr_count = len(mfr_list)
    print("メーカー数"+ str(mfr_count))
    print(mfr_list)
    
    
    try :
      mfr_1 = mfr_list[0]
    except IndexError:
      print("NoData")
      mfr_1 = ""
      
    try:  
      mfr_2 = mfr_list[1]
      
    except IndexError:
      print("NoData")
      mfr_2 = ""
      
    try:
        
      mfr_3 = mfr_list[2]
    except IndexError:
      print("NoData")  
      mfr_3 = ""
      
    try:  
      mfr_4 = mfr_list[3]
      
    except IndexError:
      print("NoDAta")
      
      mfr_4 = ""
      
    try:  
      
      mfr_5 = mfr_list[4]
      
    except IndexError:  
      
      print("NoData")
      
      mfr_5 = ""
      
    try:
        
      mfr_6 = mfr_list[5]
    except IndexError:
      print("NoData")  
      mfr_6 = ""
      
    try:
        
      mfr_7 = mfr_list[6]
      
    except IndexError:  
      print("NoData")
      mfr_7 = ""
      
    try:  
      mfr_8 = mfr_list[7]
      
    except IndexError:
      
      print("NoData")
      mfr_8 = ""
      
    try:  
      mfr_9 = mfr_list[8]
      
    except IndexError:
      print("NoData")  
      mfr_9 = ""
      
    try:  
      mfr_10 = mfr_list[9] 
    except IndexError:
      print("NoData")
      mfr_10 = ""
        


  elif week_day_type == 4 :
    
    row_no = 50
    mfr_list = []
    #金
    
    #------------------------------------------------------------------------------------------------------------------------------------
    #追加
    
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton10").click()#店間移動にチェック
    
    time.sleep(2)
    
    driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()
    
    time.sleep(5)
    
    mbs = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text
    mbs_2 = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
    
    print(mbs)
    
    if str(mbs) == '-':
      
    
      mbs_message = "◆ 【店間 - 全店合計数】\n なし"
      
    elif int(''.join(filter(str.isdigit, mbs))) > 0:
      mbs_message = "◆ 【店間 - 全店合計数】\n" + str(mbs) + "点 / " + str(mbs_2) 
      
    driver.get('http://tri.hanbai-net.com/system/21024101.aspx?id=010199')

    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02").send_keys(target_day)#開始日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02_2").send_keys(target_day)#末日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton6").click()  
      
    for n in range(0,counter):
      shop_name1 = act_ws_1.cell(row=52 + n,column=28).value
      quantity1 = act_ws_1.cell(row=52 + n,column=49).value#納品点数49
      p1 = act_ws_1.cell(row=52 + n,column=50).value#パッキン数50
      #---------------------------------------------------------------------------
      print("チェックポイント",shop_name1)
      
      #-------------------------------------------------------------------------------------------------
      

      shop_input = driver.find_element(By.ID,"ContentPlaceHolder1_DropDownListCond04")#.click()
      shop_input_select = Select(shop_input)
      shop_input_select.select_by_value(tenpo_list[shop_name1])

      driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()

      time.sleep(2)

      p_ = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
      p_f = ''.join(filter(str.isdigit, p_))

      print(p_f)

      q_ = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text
      
      if quantity1 == None:
        quantity1 = 0

      print(q_)
      print(quantity1)
      
      quantity1 = int(quantity1) + int(q_) #納品点数
      #p1 = int(p1) + int(p_f) #パッキン数
      try:
        p1 = float(p1) + float(p_f) #パッキン数
      except TypeError:
        p1 = 0
          
      
      #---------------------------------------------------------------------------
    
      if p1 == None:
        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(0) +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(0)]})
        p_list.append(row1)
      else :

        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(f"{p1: .1f}") +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(f"{p1: .1f}")]})
        p_list.append(row1)
      
    #納品メーカー名を取得  

    for c in range(0,10):
      mfr = act_ws_1[str(col_list[int(week_day_type)][int(c)]) + str(row_no)].value
   
    
      if mfr == None:
        print("No Maker")
      
      else:
        mfr_list.append(mfr)
        
        
    mfr_count = len(mfr_list)
    print("メーカー数"+ str(mfr_count))
    print(mfr_list)
    
    
    try :
      mfr_1 = mfr_list[0]
    except IndexError:
      print("NoData")
      mfr_1 = ""
      
    try:  
      mfr_2 = mfr_list[1]
      
    except IndexError:
      print("NoData")
      mfr_2 = ""
      
    try:
        
      mfr_3 = mfr_list[2]
    except IndexError:
      print("NoData")  
      mfr_3 = ""
      
    try:  
      mfr_4 = mfr_list[3]
      
    except IndexError:
      print("NoDAta")
      
      mfr_4 = ""
      
    try:  
      
      mfr_5 = mfr_list[4]
      
    except IndexError:  
      
      print("NoData")
      
      mfr_5 = ""
      
    try:
        
      mfr_6 = mfr_list[5]
    except IndexError:
      print("NoData")  
      mfr_6 = ""
      
    try:
        
      mfr_7 = mfr_list[6]
      
    except IndexError:  
      print("NoData")
      mfr_7 = ""
      
    try:  
      mfr_8 = mfr_list[7]
      
    except IndexError:
      
      print("NoData")
      mfr_8 = ""
      
    try:  
      mfr_9 = mfr_list[8]
      
    except IndexError:
      print("NoData")  
      mfr_9 = ""
      
    try:  
      mfr_10 = mfr_list[9] 
    except IndexError:
      print("NoData")
      mfr_10 = ""
          
        
  elif week_day_type == 5 :
    
    row_no = 87
    mfr_list = []
    #土
    
    
    #------------------------------------------------------------------------------------------------------------------------------------
    #追加
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton10").click()#店間移動にチェック
    
    time.sleep(2)
    
    driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()
    
    time.sleep(5)
    
    mbs = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text
    mbs_2 = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
    
    print(mbs)
    
    if str(mbs) == '-':
      
      mbs_message = str("◆ 【店間 - 全店合計数】\n" + "なし")
      
    elif int(''.join(filter(str.isdigit, mbs))) > 0:
      mbs_message = "◆ 【店間 - 全店合計数】\n" + str(mbs) + "点 / " + str(mbs_2) 
      
    driver.get('http://tri.hanbai-net.com/system/21024101.aspx?id=010199')

    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02").send_keys(target_day)#開始日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_txtCond02_2").send_keys(target_day)#末日を指定 ⇒ w_day
    driver.find_element(By.ID,"ContentPlaceHolder1_RadioButton6").click()  
      
    for n in range(0,counter):
      shop_name1 = act_ws_1.cell(row=89 + n,column=28).value
      quantity1 = act_ws_1.cell(row=89 + n,column=49).value#納品点数
      p1 = act_ws_1.cell(row=89 + n,column=50).value#パッキン数
      #---------------------------------------------------------------------------
      
      
      #------------------------------------------------------------------------------------------------------------------------------------
        
      shop_input = driver.find_element(By.ID,"ContentPlaceHolder1_DropDownListCond04")#.click()
      shop_input_select = Select(shop_input)
      shop_input_select.select_by_value(tenpo_list[shop_name1])

      driver.find_element(By.ID,"ContentPlaceHolder1_btnCondRun").click()

      time.sleep(1)

      p_ = driver.find_element(By.ID,"ContentPlaceHolder1_lblCnt").text
      p_f = ''.join(filter(str.isdigit, p_))

      print(p_f)

      q_ = driver.find_element(By.ID,"ContentPlaceHolder1_Label11").text

      print(q_)
      
      quantity1 = int(quantity1) + int(q_) #納品点数
      p1 = float(p1) + float(p_f) #パッキン数
      
      #---------------------------------------------------------------------------
    
      if p1 == None:
        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(0) +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(0)]})
        p_list.append(row1)
      else :

        print(shop_name1,'【点数】' + str(quantity1) + '点' , '【P数】' + str(f"{p1: .1f}") +'P' )
        row1 = pd.DataFrame({'店名':[shop_name1],'点数':[quantity1],'P数':[str(f"{p1: .1f}")]})
        p_list.append(row1)
      
    #納品メーカー名を取得  

    for c in range(0,10):
      mfr = act_ws_1[str(col_list[int(week_day_type)][int(c)]) + str(row_no)].value
    
    
      if mfr == None:
        print("No Maker")
      
      else:
        mfr_list.append(mfr)
        
        
    mfr_count = len(mfr_list)
    print("メーカー数"+ str(mfr_count))
    print(mfr_list)
    
    
    try :
      mfr_1 = mfr_list[0]
    except IndexError:
      print("NoData")
      mfr_1 = ""
      
    try:  
      mfr_2 = mfr_list[1]
      
    except IndexError:
      print("NoData")
      mfr_2 = ""
      
    try:
        
      mfr_3 = mfr_list[2]
    except IndexError:
      print("NoData")  
      mfr_3 = ""
      
    try:  
      mfr_4 = mfr_list[3]
      
    except IndexError:
      print("NoDAta")
      
      mfr_4 = ""
      
    try:  
      
      mfr_5 = mfr_list[4]
      
    except IndexError:  
      
      print("NoData")
      
      mfr_5 = ""
      
    try:
        
      mfr_6 = mfr_list[5]
    except IndexError:
      print("NoData")  
      mfr_6 = ""
      
    try:
        
      mfr_7 = mfr_list[6]
      
    except IndexError:  
      print("NoData")
      mfr_7 = ""
      
    try:  
      mfr_8 = mfr_list[7]
      
    except IndexError:
      
      print("NoData")
      mfr_8 = ""
      
    try:  
      mfr_9 = mfr_list[8]
      
    except IndexError:
      print("NoData")  
      mfr_9 = ""
      
    try:  
      mfr_10 = mfr_list[9] 
    except IndexError:
      print("NoData")
      mfr_10 = ""
      
    try:  
      mfr_11 = mfr_list[10] 
    except IndexError:
      print("NoData")
      mfr_11 = ""  
                

  create_list = pd.concat(p_list)
  print(create_list)

  driver.close()

  shop_1 = create_list[create_list['店名'] == "柏" ]
  shop_1_s = shop_1["店名"].values
  shop_1_q = shop_1["点数"].values
  shop_1_p = shop_1["P数"].values

  shop_2 = create_list[create_list['店名'] == "千葉" ]
  shop_2_s = shop_2["店名"].values
  shop_2_q = shop_2["点数"].values
  shop_2_p = shop_2["P数"].values

  shop_3 = create_list[create_list['店名'] == "伊勢崎" ]
  shop_3_s = shop_3["店名"].values
  shop_3_q = shop_3["点数"].values
  shop_3_p = shop_3["P数"].values


  shop_4 = create_list[create_list['店名'] == "長町" ]
  shop_4_s = shop_4["店名"].values
  # shop_4_q = shop_4["点数"].values
  # shop_4_p = shop_4["P数"].values
  
  shop_4_q = 0
  shop_4_p = 0

  shop_5 = create_list[create_list['店名'] == "TOKYO-BAY" ]
  shop_5_s = shop_5["店名"].values
  # shop_5_q = shop_5["点数"].values
  # shop_5_p = shop_5["P数"].values
  
  shop_5_q = 0
  shop_5_p = 0

  shop_6 = create_list[create_list['店名'] == "富士見" ]
  shop_6_s = shop_6["店名"].values
  shop_6_q = shop_6["点数"].values
  shop_6_p = shop_6["P数"].values

  shop_7 = create_list[create_list['店名'] == "越谷" ]
  shop_7_s = shop_7["店名"].values
  shop_7_q = shop_7["点数"].values
  shop_7_p = shop_7["P数"].values

  shop_8 = create_list[create_list['店名'] == "海老名" ]
  shop_8_s = shop_8["店名"].values
  shop_8_q = shop_8["点数"].values
  shop_8_p = shop_8["P数"].values

  shop_9 = create_list[create_list['店名'] == "むさし村山" ]
  shop_9_s = shop_9["店名"].values
  shop_9_q = shop_9["点数"].values
  shop_9_p = shop_9["P数"].values

  shop_10 = create_list[create_list['店名'] == "湘南平塚" ]
  shop_10_s = shop_10["店名"].values
  shop_10_q = shop_10["点数"].values
  shop_10_p = shop_10["P数"].values

  shop_11 = create_list[create_list['店名'] == "名取" ]
  shop_11_s = shop_11["店名"].values
  shop_11_q = shop_11["点数"].values
  shop_11_p = shop_11["P数"].values

  shop_12 = create_list[create_list['店名'] == "大高" ]
  shop_12_s = shop_12["店名"].values
  shop_12_q = shop_12["点数"].values
  shop_12_p = shop_12["P数"].values

  shop_13 = create_list[create_list['店名'] == "愛知東郷" ]
  shop_13_s = shop_13["店名"].values
  shop_13_q = shop_13["点数"].values
  shop_13_p = shop_13["P数"].values

  shop_14 = create_list[create_list['店名'] == "太田" ]
  shop_14_s = shop_14["店名"].values
  shop_14_q = shop_14["点数"].values
  shop_14_p = shop_14["P数"].values

  shop_15 = create_list[create_list['店名'] == "水戸内原" ]
  shop_15_s = shop_15["店名"].values
  shop_15_q = shop_15["点数"].values
  shop_15_p = shop_15["P数"].values 

  shop_16 = create_list[create_list['店名'] == "EXPOCITY" ]
  shop_16_s = shop_16["店名"].values
  shop_16_q = shop_16["点数"].values
  shop_16_p = shop_16["P数"].values

  shop_17 = create_list[create_list['店名'] == "川崎" ]
  shop_17_s = shop_17["店名"].values
  shop_17_q = shop_17["点数"].values
  shop_17_p = shop_17["P数"].values

  shop_18 = create_list[create_list['店名'] == "新三郷" ]
  shop_18_s = shop_18["店名"].values
  shop_18_q = shop_18["点数"].values
  shop_18_p = shop_18["P数"].values

  shop_19 = create_list[create_list['店名'] == "幕張新都心" ]
  shop_19_s = shop_19["店名"].values
  shop_19_q = shop_19["点数"].values
  shop_19_p = shop_19["P数"].values
  
  
  shop_20 = create_list[create_list['店名'] == "各務原" ]
  shop_20_s = shop_20["店名"].values
  shop_20_q = shop_20["点数"].values
  shop_20_p = shop_20["P数"].values
  
  shop_21 = create_list[create_list['店名'] == "堺" ]
  shop_21_s = shop_21["店名"].values
  shop_21_q = shop_21["点数"].values
  shop_21_p = shop_21["P数"].values
  
  
  if mfr_count == 0:
    message = "◆ 納品なし"
    
  elif mfr_count == 1 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])  
    
  elif mfr_count == 2 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])
    
  elif mfr_count == 3 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])+' ' +  str(mfr_list[2])
    
  elif mfr_count == 4 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])+' ' +  str(mfr_list[2])+' ' +  str(mfr_list[3])
    
  elif mfr_count == 5 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])+' ' +  str(mfr_list[2])+' ' +  str(mfr_list[3])+' ' +  str(mfr_list[4])
    
  elif mfr_count == 6 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])+' ' +  str(mfr_list[2])+' ' +  str(mfr_list[3])+' ' +  str(mfr_list[4])+' ' +  str(mfr_list[5]) 
    
  elif mfr_count == 7 : 
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])+' ' +  str(mfr_list[2])+' ' +  str(mfr_list[3])+' ' +  str(mfr_list[4])+' ' +  str(mfr_list[5])+' ' +  str(mfr_list[6])  
    
  elif mfr_count == 8 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])+' ' +  str(mfr_list[2])+' ' +  str(mfr_list[3])+' ' +  str(mfr_list[4])+' ' +  str(mfr_list[5])+' ' +  str(mfr_list[6])+' ' +  str(mfr_list[7])
    
  elif mfr_count == 9 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])+' ' +  str(mfr_list[2])+' ' +  str(mfr_list[3])+' ' +  str(mfr_list[4])+' ' +  str(mfr_list[5])+' ' +  str(mfr_list[6])+' ' +  str(mfr_list[7])+' ' +  str(mfr_list[8])
    
  elif mfr_count == 9 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])+' ' +  str(mfr_list[2])+' ' +  str(mfr_list[3])+' ' +  str(mfr_list[4])+' ' +  str(mfr_list[5])+' ' +  str(mfr_list[6])+' ' +  str(mfr_list[7])+' ' +  str(mfr_list[8])+' ' +  str(mfr_list[9]) 
    
  elif mfr_count == 10 :
    message = "◆【メーカー & 物流納品】\n" + str(mfr_list[0])+' ' +  str(mfr_list[1])+' ' +  str(mfr_list[2])+' ' +  str(mfr_list[3])+' ' +  str(mfr_list[4])+' ' +  str(mfr_list[5])+' ' +  str(mfr_list[6])+' ' +  str(mfr_list[7])+' ' +  str(mfr_list[8])+' ' +  str(mfr_list[9])#+' ' +  str(mfr_list[10])
    
  # print(message)  
  
  

  #w_day = '{:%Y%m%d}'.format(datetime.datetime.today())
  #year = w_day[0:4]
  #month = w_day[4:6]
  #day = int(w_day[6:8]) + 1
  
  #day = str(day).zfill(2)



  #out_file = "C:/Users/fun-f/Desktop/納品スケジュール.xlsx"
  out_file = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/シフト管理/納品スケジュール.xlsx"


  wb = pyxl.load_workbook(out_file,data_only=True)

  ws = wb["納品データ"]
  #select_day = "2022-4-14"

  select_day = str(year) + "-" + str(month) +"-" + str(day) + " 00:00:00"
  print("ここ" + str(select_day))
  for i in range(0,366):#364
    target_cell = ws["A" + str(369 + i)].value#４⇒369
    
    if str(target_cell) == str(select_day):
      print("YES")
      
      cell_r = 369 + i#４⇒369
      print(cell_r)
      
      ws["B" + str(cell_r)].value = shop_1_q[0]
      ws["C" + str(cell_r)].value = shop_1_p[0]
      
      ws["D" + str(cell_r)].value = shop_2_q[0]
      ws["E" + str(cell_r)].value = shop_2_p[0]
      
      ws["F" + str(cell_r)].value = shop_3_q[0]
      ws["G" + str(cell_r)].value = shop_3_p[0]
      
      # ws["H" + str(cell_r)].value = shop_4_q[0]
      # ws["I" + str(cell_r)].value = shop_4_p[0]
      
      # ws["J" + str(cell_r)].value = shop_5_q[0]
      # ws["K" + str(cell_r)].value = shop_5_p[0]
      
      ws["L" + str(cell_r)].value = shop_6_q[0]
      ws["M" + str(cell_r)].value = shop_6_p[0]
      
      ws["N" + str(cell_r)].value = shop_7_q[0]
      ws["O" + str(cell_r)].value = shop_7_p[0]
      
      ws["P" + str(cell_r)].value = shop_8_q[0]
      ws["Q" + str(cell_r)].value = shop_8_p[0]
      
      ws["R" + str(cell_r)].value = shop_9_q[0]
      ws["S" + str(cell_r)].value = shop_9_p[0]
      
      ws["T" + str(cell_r)].value = shop_10_q[0]
      ws["U" + str(cell_r)].value = shop_10_p[0]
      
      ws["V" + str(cell_r)].value = shop_11_q[0]
      ws["W" + str(cell_r)].value = shop_11_p[0]
      
      ws["X" + str(cell_r)].value = shop_12_q[0]
      ws["Y" + str(cell_r)].value = shop_12_p[0]
      
      ws["Z" + str(cell_r)].value = shop_13_q[0]
      ws["AA" + str(cell_r)].value = shop_13_p[0]
      
      ws["AB" + str(cell_r)].value = shop_14_q[0]
      ws["AC" + str(cell_r)].value = shop_14_p[0]
      
      ws["AD" + str(cell_r)].value = shop_15_q[0]
      ws["AE" + str(cell_r)].value = shop_15_p[0]
      
      ws["AF" + str(cell_r)].value = shop_16_q[0]
      ws["AG" + str(cell_r)].value = shop_16_p[0]
      
      ws["AH" + str(cell_r)].value = shop_17_q[0]
      ws["AI" + str(cell_r)].value = shop_17_p[0]
      
      ws["AJ" + str(cell_r)].value = shop_18_q[0]
      ws["AK" + str(cell_r)].value = shop_18_p[0]
      
      ws["AL" + str(cell_r)].value = shop_19_q[0]
      ws["AM" + str(cell_r)].value = shop_19_p[0]
      
      ws["AN" + str(cell_r)].value = shop_20_q[0]
      ws["AO" + str(cell_r)].value = shop_20_p[0]
      
      ws["AP" + str(cell_r)].value = shop_21_q[0]
      ws["AQ" + str(cell_r)].value = shop_21_p[0]
      
      
      wb.save(out_file)
      
    else:
      print(target_cell)  



  TOKEN = 'TNKXBcpEMmK4JAmRPaOyVABkA5GWIJkIHQOnsyfu4MD'#FUNの部屋トークン
  #TOKEN = 'NxPDQg0tpI4oY6BZHo7vkZ3gxPtTIijpWFyN85xL2q1'#テストの部屋トークン
  api_url = 'https://notify-api.line.me/api/notify'
  headers = {'Authorization' : 'Bearer ' + TOKEN}
  
  try :
    
    mbs_message = mbs_message
    
  except NameError:
    print("NameError")
    
    mbs_message = "◆ 【店間 - 全店合計数】\n 店間移動なし"
    
    
  
  #message = ('\n'+'柏'+'\n'+'【売上予算/実績】'+'\n' + str(mg1) +'\n' +'【P率】' +str(p1) +'\n'+ '【客数】'+ str(noc_2) +str(p2)+str(p3))
  message1 = ( "\n" + "明日のP数【 納品＋店間 】"
  '\n'+str(int(year)) + '年 ' +str(int(month)) + '月 ' + str(int(day)) + '日 ' +  '(' +str(week[week_day_type]) + ')' 
  '\n' + '\n' + str(message) + 
  '\n' + str(mbs_message) + 
  '\n'+'\n  ' +str(shop_1_s[0])+ '    ''\n【 ' +str(shop_1_q[0])+ '点 / ' + str(shop_1_p[0]) + "P 】" 
  '\n  '+str(shop_2_s[0])+ '    ''\n【 ' +str(shop_2_q[0])+ '点 / ' + str(shop_2_p[0]) + "P 】" 
  '\n  '+str(shop_3_s[0])+ '    ''\n【 ' +str(shop_3_q[0])+ '点 / ' + str(shop_3_p[0]) + "P 】" 
  # '\n  '+str(shop_4_s[0])+ '    ''\n【 ' +str(shop_4_q[0])+ '点 / ' + str(shop_4_p[0]) + "P 】"
  # '\n  '+str(shop_5_s[0])+ '    ''\n【 ' +str(shop_5_q[0])+ '点 / ' + str(shop_5_p[0]) + "P 】" 
  '\n  '+str(shop_6_s[0])+ '    ''\n【 ' +str(shop_6_q[0])+ '点 / ' + str(shop_6_p[0]) + "P 】" 
  '\n  '+str(shop_7_s[0])+ '    ''\n【 ' +str(shop_7_q[0])+ '点 / ' + str(shop_7_p[0]) + "P 】" 
  '\n  '+str(shop_8_s[0])+ '    ''\n【 ' +str(shop_8_q[0])+ '点 / ' + str(shop_8_p[0]) + "P 】" 
  '\n  '+str(shop_9_s[0])+ '    ''\n【 ' +str(shop_9_q[0])+ '点 / ' + str(shop_9_p[0]) + "P 】" 
  '\n  '+str(shop_10_s[0])+ '    ''\n【 ' +str(shop_10_q[0])+ '点 / ' + str(shop_10_p[0]) + "P 】" 
  '\n  '+str(shop_11_s[0])+ '    ''\n【 ' +str(shop_11_q[0])+ '点 / ' + str(shop_11_p[0]) + "P 】" 
  '\n  '+str(shop_12_s[0])+ '    ''\n【 ' +str(shop_12_q[0])+ '点 / ' + str(shop_12_p[0]) + "P 】" 
  '\n  '+str(shop_13_s[0])+ '    ''\n【 ' +str(shop_13_q[0])+ '点 / ' + str(shop_13_p[0]) + "P 】" 
  '\n  '+str(shop_14_s[0])+ '    ''\n【 ' +str(shop_14_q[0])+ '点 / ' + str(shop_14_p[0]) + "P 】" 
  '\n  '+str(shop_15_s[0])+ '    ''\n【 ' +str(shop_15_q[0])+ '点 / ' + str(shop_15_p[0]) + "P 】" 
  '\n  '+str(shop_16_s[0])+ '    ''\n【 ' +str(shop_16_q[0])+ '点 / ' + str(shop_16_p[0]) + "P 】" 
  '\n  '+str(shop_17_s[0])+ '    ''\n【 ' +str(shop_17_q[0])+ '点 / ' + str(shop_17_p[0]) + "P 】" 
  '\n  '+str(shop_18_s[0])+ '    ''\n【 ' +str(shop_18_q[0])+ '点 / ' + str(shop_18_p[0]) + "P 】" 
  '\n  '+str(shop_19_s[0])+ '    ''\n【 ' +str(shop_19_q[0])+ '点 / ' + str(shop_19_p[0]) + "P 】" 
  '\n  '+str(shop_20_s[0])+ '    ''\n【 ' +str(shop_20_q[0])+ '点 / ' + str(shop_20_p[0]) + "P 】" 
  '\n  '+str(shop_21_s[0])+ '    ''\n【 ' +str(shop_21_q[0])+ '点 / ' + str(shop_21_p[0]) + "P 】" 

  '\n'+'\n'+'不明点あれば古内までご連絡下さい！\n\n'

  )


  payload = {'message': message1}

  requests.post(api_url, headers=headers, params=payload)   
  print("SUCCESSFULL!!")

  #---------------------------------------------------------------------------------------
  
  #データ転記
#df_r_file = pd.DataFrame(r_file)
