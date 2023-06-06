from posixpath import splitext
from tracemalloc import stop
from PIL import Image
import pandas as pd
import openpyxl as pyxl
import os
import numpy as np
from scipy import rand

tenpo = [
    ["01001008 FUN柏","柏"],
    ["01001009 FUN千葉C-one","千葉"],
    ["01001028 FUNスマーク伊勢崎","伊勢崎"],
    # ["01001032 FUNララガーデン長町","長町"],
    # ["01001033 FUNららぽーとTOKYO-BAY","船橋"],
    ["01001034 FUNららぽーと富士見","富士見"],
    ["01001036 FUNイオンレイクタウン","レイク"],
    ["01001038 FUNららぽーと海老名","海老名"],
    ["01001039 FUNイオンモールむさし村山","むさし"],
    ["01001040 FUNららぽーと湘南平塚","平塚"],
    ["01001041 FUNイオンモール名取","名取"],
    ["01001042 FUNイオンモール大高","大高"],
    ["01001043 FUNららぽーと愛知東郷","東郷町"],
    ["01001044 FUNイオンモール太田","太田"],
    ["01001045 FUNイオンモール水戸内原","水戸"],
    ["01001046 FUNららぽーとEXPOCITY","EXPO"],
    ["01001047 FUNラゾーナ川崎プラザ","川崎"],
    ["01001048 FUNららぽーと新三郷","新三郷"],
    ["01001049 FUNイオンモール幕張新都心","幕張"],
    ["01001050 FUNイオンモール各務原","各務原"],
    ["01001051 FUNららぽーと堺","堺"],
    
]

dr_files = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/data_folder'#今週実績
dr_files2 = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/previous_data'#過去実績
dr_files3 = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/sales_values'#売上集計

dr_read = os.listdir(dr_files)
dr_read2 = os.listdir(dr_files2)
dr_read3 = os.listdir(dr_files3)



#out_put_file = "C:/Users/fun-f/Desktop/analysis/画像テスト.xlsx"
selectfile = os.listdir("C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/create_file/")
out_put_file = os.path.join("C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/create_file/",selectfile[0])

img_stock = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/item_image_stock"
img_stock_list = os.listdir(img_stock)

#入荷情報リストを作成
#<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<

#フォルダーパス
new_arrival_path = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/new_arrival/"

new_arrival_file = os.listdir(new_arrival_path)
print(new_arrival_file)

col_list = [
  "B","F","J","N","R","V"
]

wb = pyxl.load_workbook(os.path.join(new_arrival_path,new_arrival_file[1]))
#wb = pyxl.load_workbook("C:/Users/fun-f/Desktop/analysis/new_arrival/入荷予定MAP_20221115183651_1.xlsx")

sheet_name = wb.sheetnames
#sheet_count = len(sheet_name)

project_list = []
index_num = [16,39,62,85,108,131,154]#85,108,131

for sheet_n in sheet_name:
  ws = wb[sheet_n]
  
  class Items:
    
    def __init__(self, name, item_cd, category_cd , producttion_number ,size):
      
      #商品名
      self.name = name
      #商品CD
      self.item_cd = item_cd
      #アイテムCD
      self.category_cd = category_cd
      #生産枚数
      self.producttion_number = producttion_number
      #サイズ
      self.size = size
      
  for index_x in index_num:    
      
    for c_no in range(0,6):
      print(c_no)  
      Items.name = ws[col_list[c_no] + str(index_x + 1)].value
      Items.item_cd = ws[col_list[c_no] + str(index_x)].value
      Items.category_cd = str(ws[col_list[c_no] + str(index_x)].value)[2:4]
      Items.producttion_number = ws[col_list[c_no] + str(index_x + 2)].value
      Items.size = ws[col_list[c_no] + str(index_x + 4)].value
      
      data_RC = pd.DataFrame({"商品名":[Items.name],"商品CD":[Items.item_cd],"アイテムCD":[Items.category_cd],"生産枚数":[Items.producttion_number],"サイズ":[Items.size]})
      
      if Items.name == None :
        print("None")
        
      else:  
        project_list.append(data_RC)
  
new_arrival_list = pd.DataFrame(pd.concat(project_list,axis=0))
#print(new_arrival_list)
sort_list = new_arrival_list.sort_values("生産枚数",ascending=False)

item_cd_uniq = np.unique(sort_list["アイテムCD"].values)
print(item_cd_uniq)

#カテゴリー生産総枚数順に優先順位を設定

production_rank_list = []
for item_n in item_cd_uniq:
  key_item = sort_list[sort_list["アイテムCD"] == item_n]
  category_number = sum(key_item["生産枚数"].values)
  print(category_number)
  
  data_index = pd.DataFrame({"アイテムCD":[item_n],"生産枚数":[category_number]})
  production_rank_list.append(data_index)
  
ranking_list = pd.concat(production_rank_list,axis=0).sort_values("生産枚数",ascending=False).head(4)  

print(ranking_list)

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

item_category_dic = {
          "01":"OP",
          "02":"CD",
          "03":"JK",
          "04":"KT",
          "05":"CS",
          "06":"CT",
          "07":"BL",
          "08":"SK",
          "09":"PT",
          "10":"TR",
          "11":"INN",
          "12":"SETUP",
          "13":"ACC",
          "15":"SHOES",
        }

img_columns1_1 = ["S45"]
img_columns2_1 = ["AK45"]
img_columns3_1 = ["S57"]
img_columns4_1 = ["AK57"]

img_theme = [img_columns1_1,img_columns2_1,img_columns3_1,img_columns4_1]

img_columns1 = ["U","Y","AC","AG",]
img_columns2 = ["AN","AR","AV","AZ"]

img_columns3 = ["U","Y","AC","AG",]
img_columns4 = ["AN","AR","AV","AZ"]
          
img_columns = [img_columns1,img_columns2,img_columns3,img_columns4]          

img_index1 = [42,43,44,52,53]
img_index2 = [42,43,44,52,53]
img_index3 = [54,55,56,65,66]
img_index4 = [54,55,56,65,66]

img_index_list = [img_index1,img_index2,img_index3,img_index4]

best_col = ["T","W","Z","AC","AF","AI","AL","AO","AR","AU","AX","BA"]

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#店別ベスト作成

data_concat_list = []
  
for select_i in tenpo:
  
  shop_i = select_i
    
    
  week2 = pd.read_csv('C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/data_folder/' + str(shop_i[1]) + ".csv",encoding='cp932')#ok
  
  df_week1 = pd.DataFrame(week2)

  print(df_week1)

  item_cd = pd.DataFrame(df_week1["商品コード"].astype('str').str.zfill(10).values,columns=["商品CD"])
  item_name = pd.DataFrame(df_week1["商品名"].values,columns=["商品名"])
  category_cd = pd.DataFrame(df_week1["商品コード"].astype('str').str.zfill(10).str[2:4].values,columns=["アイテムCD"])
  quantity = pd.DataFrame(df_week1['合計数量'].values,columns=["数量"])
  amount = pd.DataFrame(df_week1['合計金額'].values,columns=["金額"])
  #shop_name = pd.DataFrame([shop_i[2]],columns=["店舗"])

  df_week1_values = pd.concat([item_cd,item_name,category_cd,quantity,amount],axis=1)

  filter1_df_week1_values = df_week1_values[df_week1_values["アイテムCD"] != "98" ]

  filter2_df_week1_values = filter1_df_week1_values[(filter1_df_week1_values["商品名"] != "ｷﾚｲﾏｽｸ") & (filter1_df_week1_values["商品名"] != "ｻﾝﾌﾟﾙ") ]
  
  filter2_df_week1_values2 = pd.DataFrame(filter2_df_week1_values)
  
  sort_filter2_df_week1_values2 = filter2_df_week1_values2.sort_values("数量",ascending=False)#★★★ 修正 ★★★
  #sort_filter2_df_week1_values3 = filter2_df_week1_values2.sort_values("金額",ascending=False)#★★★ 修正 ★★★
  
  for low_data in sort_filter2_df_week1_values2.values :#★★★ 修正 ★★★
    item_cd = pd.DataFrame([low_data[0]],columns=["商品CD"])
    item_name = pd.DataFrame([low_data[1]],columns=["商品名"])
    category_cd = pd.DataFrame([low_data[2]],columns=["アイテムCD"])
    quantity = pd.DataFrame([low_data[3]],columns=["数量"])
    amount = pd.DataFrame([low_data[4]],columns=["金額"])
    shop_name = pd.DataFrame([shop_i[1]],columns=["店舗"])
    
    low_data2 = pd.concat([item_cd,item_name,category_cd,quantity,amount,shop_name],axis=1)
    
    data_concat_list.append(low_data2)
    

  all_amount = sum(filter2_df_week1_values["金額"].values)

  op_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "01"]
  op_1_amount = sum(op_1["金額"].values)
  try :
    op_1_ratio = op_1_amount / all_amount 
    
  except ZeroDivisionError:
    
    op_1_ratio = 0
    
  op_list = [op_1_amount,op_1_ratio]

  cd_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "02"]
  cd_1_amount = sum(cd_1["金額"].values)
  
  try:
    cd_1_ratio = cd_1_amount / all_amount 
    
  except ZeroDivisionError:
    cd_1_ratio = 0
    
  cd_list = [cd_1_amount,cd_1_ratio]
  
  
  jk_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "03"]
  jk_1_amount = sum(jk_1["金額"].values)
  try:
    jk_1_ratio = jk_1_amount / all_amount 
    
  except ZeroDivisionError:
    jk_1_ratio = 0
    
  jk_list = [jk_1_amount,jk_1_ratio]

  kt_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "04"]
  kt_1_amount = sum(kt_1["金額"].values)
  
  try :
    kt_1_ratio = kt_1_amount / all_amount 
    
  except ZeroDivisionError:
    kt_1_ratio = 0
    
  kt_list = [kt_1_amount,kt_1_ratio]

  cs_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "05"]
  cs_1_amount = sum(cs_1["金額"].values)
  
  try :
    cs_1_ratio = cs_1_amount / all_amount 
    
  except ZeroDivisionError:
    cs_1_ratio = 0
    
  cs_list = [cs_1_amount,cs_1_ratio]

  ct_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "06"]
  ct_1_amount = sum(ct_1["金額"].values)
  
  try:
    ct_1_ratio = ct_1_amount / all_amount 
    
  except ZeroDivisionError:
    ct_1_ratio = 0
    
  ct_list = [ct_1_amount,ct_1_ratio]

  bl_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "07"]
  bl_1_amount = sum(bl_1["金額"].values)
  
  try :
    bl_1_ratio = bl_1_amount / all_amount 
    
  except ZeroDivisionError :
    bl_1_ratio = 0
    
  bl_list = [bl_1_amount,bl_1_ratio]

  sk_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "08"]
  sk_1_amount = sum(sk_1["金額"].values)
  
  try :
    sk_1_ratio = sk_1_amount / all_amount 
    
  except ZeroDivisionError:
    sk_1_ratio = 0
    
  sk_list = [sk_1_amount,sk_1_ratio]

  pt_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "09"]
  pt_1_amount = sum(pt_1["金額"].values)
  
  try :
    pt_1_ratio = pt_1_amount / all_amount 
    
  except ZeroDivisionError:
    pt_1_ratio = 0
    
  pt_list = [pt_1_amount,pt_1_ratio]

  tr_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "10"]
  tr_1_amount = sum(tr_1["金額"].values)
  
  try :
    tr_1_ratio = tr_1_amount / all_amount 
    
  except ZeroDivisionError:
    
    tr_1_ratio = 0
    
  tr_list = [tr_1_amount,tr_1_ratio]

  inn_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "11"]
  inn_1_amount = sum(inn_1["金額"].values)
  
  try:
    inn_1_ratio = inn_1_amount / all_amount 
    
  except ZeroDivisionError:
    inn_1_ratio = 0
    
  inn_list = [inn_1_amount,inn_1_ratio]

  setup_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "12"]
  setup_1_amount = sum(setup_1["金額"].values)
  
  try:
    setup_1_ratio = setup_1_amount / all_amount
    
  except ZeroDivisionError:
    setup_1_ratio = 0
    
  setup_list = [setup_1_amount,setup_1_ratio]

  acc_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "13"]
  acc_1_amount = sum(acc_1["金額"].values)
  
  try:
    acc_1_ratio = acc_1_amount / all_amount
    
  except ZeroDivisionError:
    
    acc_1_ratio = 0
    
  acc_list = [acc_1_amount,acc_1_ratio]
  
  
  sh_1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == "15"]
  sh_1_amount = sum(sh_1["金額"].values)
  
  try:
    sh_1_ratio = sh_1_amount / all_amount
    
  except ZeroDivisionError:
    
    sh_1_ratio = 0
    
  sh_list = [sh_1_amount,sh_1_ratio]

  out_put_list = [
    
    op_list,
    cd_list,
    jk_list,
    kt_list,
    cs_list,
    ct_list,
    bl_list,
    sk_list,
    pt_list,
    tr_list,
    inn_list,
    setup_list,
    acc_list,
    sh_list
    
  ]
  
  print(all_amount)


  
print(data_concat_list)  

#data_concat_list2 = pd.DataFrame([data_concat_list])

try :

  df_data_concat_list = pd.concat(data_concat_list,axis=0)
  
except ValueError:  
  
  item_cd = pd.DataFrame([""],columns=["商品CD"])
  item_name = pd.DataFrame([""],columns=["商品名"])
  category_cd = pd.DataFrame([""],columns=["アイテムCD"])
  quantity = pd.DataFrame([0],columns=["数量"])
  amount = pd.DataFrame([0],columns=["金額"])
  shop_name = pd.DataFrame([""],columns=["店舗"])
  
  low_data2 = pd.concat([item_cd,item_name,category_cd,quantity,amount,shop_name],axis=1)
  
  data_concat_list.append(low_data2)
  
  
  
  df_data_concat_list = pd.concat(data_concat_list,axis=0)
  
  
  
print(df_data_concat_list)



#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
#追加ここまで

wb_out = pyxl.load_workbook(out_put_file)

for sheet_name in tenpo:
  ws_out = wb_out[sheet_name[1]]  

  best_key = df_data_concat_list[df_data_concat_list["店舗"] == sheet_name[1]].sort_values("金額",ascending=False).head(12)

  for best_n,no_ in zip(best_key.values,range(0,13)):
    best_target = best_n[0]
    i_name = best_n[1]
    i_sales = best_n[4]
    ws_out[best_col[no_] + str(4)] = i_name
    ws_out[best_col[no_] + str(5)] = i_sales
    
    
    for img_x in img_stock_list:
        base,ext = splitext(img_x)
        
        if str(base) == str(best_target) :
          
          t_path = img_stock + "/" + img_x

          img = Image.open(t_path)
          
          re_img = img.resize((420, 600))
          
          re_img.save(t_path)
          
          best_pasting_img = pyxl.drawing.image.Image(t_path)
      
          best_pasting_img.anchor = str(best_col[no_] + str(6))#画像挿入
          
          ws_out.add_image(best_pasting_img)
    

  for rank_n,counter in zip(ranking_list.values,range(0,4)):

    
    key_2 = sort_list[sort_list["アイテムCD"] == rank_n[0]].head(4)
    counter_2 = len(key_2["商品名"].values)
    print(counter_2)
    ws_out[img_theme[counter][0]].value = item_category_dic[rank_n[0]]
    
    row_counter = 0
    for out_data in key_2.values:
      
      print(out_data)
      ws_out[img_columns[counter][row_counter ] + str(img_index_list[counter][0])].value = out_data[1]#品番
      ws_out[img_columns[counter][row_counter ] + str(img_index_list[counter][1])].value = out_data[0]#商品名
      ws_out[img_columns[counter][row_counter ] + str(img_index_list[counter][4])].value = out_data[3]#生産枚数
      ws_out[img_columns[counter][row_counter ] + str(img_index_list[counter][3])].value = out_data[4]#生産枚数
      target = out_data[1]
      

      for img_x in img_stock_list:
        base,ext = splitext(img_x)
        
        if str(base) == str(target) :
          
          t_path = img_stock + "/" + img_x

          img = Image.open(t_path)
          
          re_img = img.resize((560, 750))
          
          re_img.save(t_path)
          
          pasting_img = pyxl.drawing.image.Image(t_path)
      
          pasting_img.anchor = str(img_columns[counter][row_counter ] + str(img_index_list[counter][2]))#画像挿入
          
          ws_out.add_image(pasting_img)
    
      row_counter += 1
  
wb_out.save(out_put_file)
#wb_out.save(os.path.join("C:/Users/fun-f/Desktop/analysis","週間分析1.xlsx"))  
  
  
#<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<<
  
  
  


#img_stock = "C:/Users/fun-f/Desktop/analysis/item_image_stock"
#img_stock_list = os.listdir(img_stock)

#target = "1105050180"

#for img_x in img_stock_list:
  #base,ext = splitext(img_x)
  #print(img_x)
  
  #if base == target :
    #print("ターゲット",img_x)
    #t_path = img_stock + "/" + img_x

    #img = Image.open(t_path)
    

    
    #width = 20#19.32
    
    #height = int(img.height * width / img.width)
    
    #re_img = img.resize((width, height))
    #re_img = img.resize((610, 780))
    
    #re_img.save(t_path)
    
    #pasting_img = pyxl.drawing.image.Image(t_path)
    #pasting_img.anchor = "B8"
    #pasting_img.anchor = "U56"
    
    #ws_out.add_image(pasting_img)
    
#wb_out.save(os.path.join("C:/Users/fun-f/Desktop/analysis","週間分析1.xlsx"))