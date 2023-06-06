#日別予算を出力するプログラムになります。

import openpyxl as pyxl
import pandas as pd
import os
import numpy as np
from decimal import *
import time


tenpo = [
    ["01001008 FUN柏","柏"],
    ["01001009 FUN千葉C-one","千葉"],
    ["01001028 FUNスマーク伊勢崎","伊勢崎"],
    #["01001032 FUNララガーデン長町","長町"],
    #["01001033 FUNららぽーとTOKYO-BAY","船橋"],
    ["01001034 FUNららぽーと富士見","富士見"],
    ["01001036 FUNイオンレイクタウン","レイク"],
    ["01001038 FUNららぽーと海老名","海老名"],
    ["01001039 FUNイオンモールむさし村山","むさし"],
    ["01001040 FUNららぽーと湘南平塚","平塚"],
    ["01001041 FUNイオンモール名取","名取"],
    ["01001042 FUNイオンモール大高","大高"],
    ["01001043 FUNららぽーと愛知東郷","東郷町"],
    ["01001044 FUNイオンモール太田","太田"],
    ["01001045 FUNイオンモール水戸内原","水戸"],
    ["01001046 FUNららぽーとEXPOCITY","EXPO"],
    ["01001047 FUNラゾーナ川崎プラザ","川崎"],
    ["01001048 FUNららぽーと新三郷","新三郷"],
    ["01001049 FUNイオンモール幕張新都心","幕張"],
    ["01001050 FUNイオンモール各務原","各務原"],
    ["01001051 FUNららぽーと堺","堺"],
    
]

buget_col = ["U","Z","AE","AJ","AO","AT","AY"]

#臨時パス
path = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/create_file/'

file = os.listdir(path)

path = os.path.join(path,file[0])
#'東郷町日別予算設定管理.csv'
buget_folder = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/buget'


files = os.listdir(buget_folder)

###########################################################################################
wb = pyxl.load_workbook(path)

      #ws = wb[str(key_1[1])]
      
for sheet_name in tenpo:    
  
  file_name = sheet_name[1] + '日別予算設定管理.csv'
      
  ws = wb[sheet_name[1]]
  r_file = pd.read_csv(os.path.join(buget_folder,file_name),encoding="cp932")

  print(r_file)


  df_r_file = pd.DataFrame(r_file)

  buget = df_r_file["売上予算"].values
  
  #週計予算
  all_buget = sum(buget)
  
  #売上構成比の合計値(※靴を除く)
  #在庫金額￥20000以下の場合は除外
  ele1 = []
  for index_n in range(17,30):
    ele2 = ws["G" + str(index_n)].value
    ele3 = ws["E" + str(index_n)].value
    if int(ele2) >= 10 :#20000
      ele1.append(ele3)
      
  sum_ele = sum(ele1)#合計

  #日割予算を入力
  for index_n in range(0,7):
    
    ws[str(buget_col[int(index_n)]) + str(11)].value = buget[int(index_n)]

    
    #アイテム予算を入力
    for index_n2 in range(17,30):
      element_inv =  ws["G" + str(index_n2)].value
      element_ratio = ws["L" + str(index_n2)].value
      
      if int(element_inv) >= 10 :#20000
        print(Decimal(str(element_ratio / sum_ele)).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP))
        
        #日割予算
        ws[str(buget_col[index_n]) + str(index_n2)].value = buget[index_n] * Decimal(str(element_ratio / sum_ele)).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)
        #週計予算
        ws["R" + str(index_n2)].value = all_buget * Decimal(str(element_ratio / sum_ele)).quantize(Decimal("0.01"),rounding=ROUND_HALF_UP)
        
      elif int(element_inv) < 10 :#20000
        
        #日割予算
        ws[str(buget_col[index_n]) + str(index_n2)].value = 0
        
        #週計予算
        ws["R" + str(index_n2)].value = 0
        
      
    #OP/SETUP
    ws[str(buget_col[index_n]) + str(33)].value = ws[buget_col[index_n] + str(17)].value + ws[buget_col[index_n] + str(28)].value
    
    ws["R" + str(33)].value = ws["R" + str(17)].value + ws["R" + str(28)].value
    
    #TOPS
    ws[str(buget_col[index_n]) + str(34)].value = ws[buget_col[index_n] + str(20)].value + ws[buget_col[index_n] + str(21)].value + ws[buget_col[index_n] + str(23)].value + ws[buget_col[index_n] + str(26)].value
    
    ws["R" + str(34)].value = ws["R" + str(20)].value + ws["R" + str(21)].value + ws["R" + str(23)].value + ws["R" + str(26)].value
    
    #BOTOMMS
    ws[str(buget_col[index_n]) + str(35)].value = ws[buget_col[index_n] + str(24)].value + ws[buget_col[index_n] + str(25)].value
    
    ws["R" + str(35)].value = ws["R" + str(24)].value + ws["R" + str(25)].value
    
    #羽織
    ws[str(buget_col[index_n]) + str(36)].value = ws[buget_col[index_n] + str(18)].value + ws[buget_col[index_n] + str(19)].value + ws[buget_col[index_n] + str(22)].value
    
    ws["R" + str(36)].value = ws["R" + str(18)].value + ws["R" + str(19)].value + ws["R" + str(22)].value
    
    #INNER
    ws[str(buget_col[index_n]) + str(37)].value = ws[buget_col[index_n] + str(27)].value
    
    ws["R" + str(37)].value = ws[ "R" + str(27)].value
    
    #ACC
    ws[str(buget_col[index_n]) + str(38)].value = ws[buget_col[index_n] + str(29)].value
    ws["R" + str(38)].value = ws["R" + str(29)].value
    
    #合計出力
    
    ws["R" + str(30)].value = all_buget
    ws["R" + str(39)].value = all_buget
    
    ws[str(buget_col[index_n]) + str(30)].value = ws[buget_col[index_n] + str(17)].value + ws[buget_col[index_n] + str(28)].value + ws[buget_col[index_n] + str(20)].value + ws[buget_col[index_n] + str(21)].value + ws[buget_col[index_n] + str(23)].value + ws[buget_col[index_n] + str(26)].value + ws[buget_col[index_n] + str(24)].value + ws[buget_col[index_n] + str(25)].value + ws[buget_col[index_n] + str(18)].value + ws[buget_col[index_n] + str(19)].value + ws[buget_col[index_n] + str(22)].value + ws[buget_col[index_n] + str(27)].value + ws[buget_col[index_n] + str(29)].value
    
    
    ws[str(buget_col[index_n]) + str(39)].value = ws[buget_col[index_n] + str(17)].value + ws[buget_col[index_n] + str(28)].value + ws[buget_col[index_n] + str(20)].value + ws[buget_col[index_n] + str(21)].value + ws[buget_col[index_n] + str(23)].value + ws[buget_col[index_n] + str(26)].value + ws[buget_col[index_n] + str(24)].value + ws[buget_col[index_n] + str(25)].value + ws[buget_col[index_n] + str(18)].value + ws[buget_col[index_n] + str(19)].value + ws[buget_col[index_n] + str(22)].value + ws[buget_col[index_n] + str(27)].value + ws[buget_col[index_n] + str(29)].value
    
    
    
      
  time.sleep(2)  
  wb.save(os.path.join(path,file[0]))   
    
    
  #####################################################################################################