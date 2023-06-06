import pandas as pd
import openpyxl as pyxl
import os
import zipfile
import shutil
from pathlib import Path

from openpyxl.drawing.image import Image

#フォルダーパス
new_arrival_path = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/new_arrival/"


new_arrival_file = os.listdir(new_arrival_path)
print(new_arrival_file[1])

col_list = [
  "B","F","J","N","R","V"
]

wb = pyxl.load_workbook(os.path.join(new_arrival_path,new_arrival_file[1]))
#wb = pyxl.load_workbook("C:/Users/fun-f/Desktop/analysis/new_arrival/入荷予定MAP_20221115183758_1.xlsx")


sheet_name = wb.sheetnames
#sheet_count = len(sheet_name)

for sheet_n in sheet_name:
  ws = wb[sheet_n]
  
  class Items:
    
    def __init__(self,name1,name2,name3,name4,name5,name6,name7,name8,name9,name10,name11,name12,name13,name14,name15,name16,name17,name18,name19,name20,name21,name22,name23,name24,name25,name26,name27,name28,name29,name30,name31,name32,name33,name34,name35,name36 ):
      
      self.name1 = name1
      self.name2 = name2
      self.name3 = name3
      self.name4 = name4
      self.name5 = name5
      self.name6 = name6
      
      self.name7 = name7
      self.name8 = name8
      self.name9 = name9
      self.name10 = name10
      self.name11 = name11
      self.name12 = name12
      
      self.name13 = name13
      self.name14 = name14
      self.name15 = name15
      self.name16 = name16
      self.name17 = name17
      self.name18 = name18
      
      self.name19 = name19
      self.name20 = name20
      self.name21 = name21
      self.name22 = name22
      self.name23 = name23
      self.name24 = name24
      
      self.name25 = name25
      self.name26 = name26
      self.name27 = name27
      self.name28 = name28
      self.name29 = name29
      self.name30 = name30
      
      self.name31 = name31
      self.name32 = name32
      self.name33 = name33
      self.name34 = name34
      self.name35 = name35
      self.name36 = name36
      
      
  index_no1 = 16
  
  Items.name1 = ws[col_list[0] + str(index_no1)].value
  Items.name2 = ws[col_list[1] + str(index_no1)].value
  Items.name3 = ws[col_list[2] + str(index_no1)].value
  Items.name4 = ws[col_list[3] + str(index_no1)].value
  Items.name5 = ws[col_list[4] + str(index_no1)].value
  Items.name6 = ws[col_list[5] + str(index_no1)].value
  
  index_no2 = 39
  
  Items.name7 = ws[col_list[0] + str(index_no2)].value
  Items.name8 = ws[col_list[1] + str(index_no2)].value
  Items.name9 = ws[col_list[2] + str(index_no2)].value
  Items.name10 = ws[col_list[3] + str(index_no2)].value
  Items.name11 = ws[col_list[4] + str(index_no2)].value
  Items.name12 = ws[col_list[5] + str(index_no2)].value
  
  index_no3 = 62
  
  Items.name13 = ws[col_list[0] + str(index_no3)].value
  Items.name14 = ws[col_list[1] + str(index_no3)].value
  Items.name15 = ws[col_list[2] + str(index_no3)].value
  Items.name16 = ws[col_list[3] + str(index_no3)].value
  Items.name17 = ws[col_list[4] + str(index_no3)].value
  Items.name18 = ws[col_list[5] + str(index_no3)].value
  
  index_no4 = 85
  
  Items.name19 = ws[col_list[0] + str(index_no4)].value
  Items.name20 = ws[col_list[1] + str(index_no4)].value
  Items.name21 = ws[col_list[2] + str(index_no4)].value
  Items.name22 = ws[col_list[3] + str(index_no4)].value
  Items.name23 = ws[col_list[4] + str(index_no4)].value
  Items.name24 = ws[col_list[5] + str(index_no4)].value
  
  index_no5 = 108
  
  Items.name25 = ws[col_list[0] + str(index_no5)].value
  Items.name26 = ws[col_list[1] + str(index_no5)].value
  Items.name27 = ws[col_list[2] + str(index_no5)].value
  Items.name28 = ws[col_list[3] + str(index_no5)].value
  Items.name29 = ws[col_list[4] + str(index_no5)].value
  Items.name30 = ws[col_list[5] + str(index_no5)].value
  
  
  index_no6 = 131
  
  Items.name31 = ws[col_list[0] + str(index_no6)].value
  Items.name32 = ws[col_list[1] + str(index_no6)].value
  Items.name33 = ws[col_list[2] + str(index_no6)].value
  Items.name34 = ws[col_list[3] + str(index_no6)].value
  Items.name35 = ws[col_list[4] + str(index_no6)].value
  Items.name36 = ws[col_list[5] + str(index_no6)].value
  
  index_no7 = 154
  
  Items.name37 = ws[col_list[0] + str(index_no7)].value
  Items.name38 = ws[col_list[1] + str(index_no7)].value
  Items.name39 = ws[col_list[2] + str(index_no7)].value
  Items.name40 = ws[col_list[3] + str(index_no7)].value
  Items.name41 = ws[col_list[4] + str(index_no7)].value
  Items.name42 = ws[col_list[5] + str(index_no7)].value
  
  
      
      
  item_list = [Items.name1,Items.name2,Items.name3,Items.name4,Items.name5,Items.name6,Items.name7,Items.name8,Items.name9,Items.name10,Items.name11,Items.name12,Items.name13,Items.name14,Items.name15,Items.name16,Items.name17,Items.name18,Items.name19,Items.name20,Items.name21,Items.name22,Items.name23,Items.name24,Items.name25,Items.name26,Items.name27,Items.name28,Items.name29,Items.name30,Items.name31,Items.name32,Items.name33,Items.name34,Items.name35,Items.name36,Items.name37,Items.name38,Items.name39,Items.name40,Items.name41,Items.name42]
  
  item_count = list(filter(lambda x: x != None, item_list))
  
  xlsx_zip = zipfile.ZipFile(os.path.join(new_arrival_path,new_arrival_file[1]))
  zip_files = xlsx_zip.namelist()
  images_path = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/images'
  
  file_no = 1

  for images_n in zip_files:
    #print(images_n)
    #if images_n.endswith(".jpeg") == True:
    if images_n.startswith("xl/media/"):
      print(images_n)
      
      #shutil.move(Path(images_n).name ,images_path)
      img_file = xlsx_zip.open(images_n)
      img_bytes = img_file.read()
      
      xlsx_path = Path(os.path.join("analysis/new_arrival",new_arrival_file[1]))
      print(xlsx_path)

      #print(img_bytes)
      
      #img_path = "images" / (xlsx_path.stem + "_" + Path(images_n).name)

      
      #with xlsx_zip.open(images_n,mode="w") as f:
      #images_path.write(img_bytes)
        
      #img_file.close()  
      
  #xlsx_zip.close()    
        
import zipfile
from pathlib import Path

# 画像を取り出すExcelブックのパス
#xlsx_path = Path("xl_files/sample1.xlsx")
xlsx_path = Path(os.path.join(new_arrival_path,new_arrival_file[1]))



# zipfileモジュールでExcelブックを開く
xlsx_zip = zipfile.ZipFile(xlsx_path)
zipped_files = xlsx_zip.namelist()

# 画像を保存するフォルダー
#img_dir = Path("xl_images")
img_dir = Path("C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/images")
img_dir.mkdir(exist_ok=True)

# xlsxファイルの中身を1つずつループ
item_no = 0
for file in zipped_files:
    if file.startswith("xl/media/"):
        # 画像ファイルを開く
        img_file = xlsx_zip.open(file)
        # 画像ファイルの読み込み
        img_bytes = img_file.read()

        # 保存する画像ファイル名には、「xlsxファイル名_」を先頭に付ける
        #img_path = img_dir / (xlsx_path.stem + "_" + Path(file).name)
        img_path = img_dir /( Path(file).name)
        # 画像ファイルの保存
        with img_path.open(mode="wb") as f:
            f.write(img_bytes)
            print(f)
        img_file.close()
  

xlsx_zip.close()  

import time
#from PIL import Image

#>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>  

time.sleep(2)

out_put_file = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/stock/画像テスト.xlsx"

img_dir = "C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/images"

out_wb = pyxl.load_workbook(out_put_file)

out_ws = out_wb["Sheet1"]

files = os.listdir(img_dir)

for image_file in files:
  print(image_file)
  
  base,ext = os.path.splitext(image_file)
  
  print(ext)
  
  if str(image_file) == "image1" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name1)  + ext))
    
  elif str(image_file) == "image2" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name2)  + ext))  
    
  elif str(image_file) == "image3" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name3)  + ext))  
    
  elif str(image_file) == "image4" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name4)  + ext))  
    
  elif str(image_file) == "image5" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name5)  + ext))  
    
  elif str(image_file) == "image6" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name6)  + ext))  
    
  elif str(image_file) == "image7" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name7)  + ext))  
    
    
  elif str(image_file) == "image8" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name8)  + ext))  
    
    
  elif str(image_file) == "image9" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name9)  + ext))  
    
  elif str(image_file) == "image10" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name10)  + ext))  
    
  elif str(image_file) == "image11" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name11)  + ext))  
    
  elif str(image_file) == "image12" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name12)  + ext))  
    
    
  elif str(image_file) == "image13" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name13)  + ext))  
    
  elif str(image_file) == "image14" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name14)  + ext))  
    
  elif str(image_file) == "image15" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name15)  + ext))  
    
  elif str(image_file) == "image16" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name16)  + ext))  
  
  elif str(image_file) == "image17" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name17)  + ext))  
    
  elif str(image_file) == "image18" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name18)  + ext))  
    
    
  elif str(image_file) == "image19" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name19)  + ext))  
    
  elif str(image_file) == "image20" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name20)  + ext))  
    
  elif str(image_file) == "image21" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name21)  + ext))  
    
    
  elif str(image_file) == "image22" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name22)  + ext))  
    
  elif str(image_file) == "image23" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name23)  + ext))  
    
  elif str(image_file) == "image24" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name24)  + ext))  
    
  elif str(image_file) == "image25" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name25)  + ext))  
  
  elif str(image_file) == "image26" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name26)  + ext))  
    
  elif str(image_file) == "image27" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name27)  + ext))    
    
  elif str(image_file) == "image28" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name28)  + ext))  
    
  elif str(image_file) == "image29" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name29)  + ext))  
    
  elif str(image_file) == "image30" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name30)  + ext))  
    
    
  elif str(image_file) == "image31" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name31)  + ext))  
    
  elif str(image_file) == "image32" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name32)  + ext))  
    
  elif str(image_file) == "image33" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name33)  + ext))  
    
  elif str(image_file) == "image34" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name34)  + ext))  
  
  elif str(image_file) == "image35" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name35)  + ext))  
    
  elif str(image_file) == "image36" + ext:
               
    os.rename(os.path.join(img_dir,image_file), os.path.join(img_dir,str(Items.name36)  + ext))    
    
  else:
    print("No File")  
        
rename_files = os.listdir(img_dir)    

for file_name in rename_files:
  print("ここ",file_name)
  
  try:
    shutil.move(os.path.join(img_dir,file_name),"C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/item_image_stock")
  except shutil.Error:
    
    os.remove(os.path.join(img_dir,file_name))
    print("同名ファイルがある為削除しました")
      


    
  
  #image_select = Image(img_dir + "/" + image_file)
  #img = Image.open(image_select)
  
  #img.resize((300,200))

  #out_ws.add_image(image_select,"B3")

#out_wb.save(out_put_file)




  
      
      
      
    
 
