import pandas as pd
import openpyxl as pyxl

import numpy as np
import os

from .parts_31 import create_filename

week1 = pd.read_csv('C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/data_folder/全店.csv',encoding='cp932')#今週実績
week1_sales_values = pd.read_csv('C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/sales_values/売上実績4.csv',encoding='cp932')#今週売上集計


previous_week1 = pd.read_csv("C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/previous_data/全店1.csv",encoding='cp932')#過去実績今週
previous_week1_sales_values = pd.read_csv('C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/sales_values/売上実績1.csv',encoding='cp932')#前週売上集計

previous_week2 = pd.read_csv("C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/previous_data/全店2.csv",encoding='cp932')#過去実績翌週
previous_week2_sales_values = pd.read_csv('C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/sales_values/売上実績2.csv',encoding='cp932')#今週売上集計

previous_week3 = pd.read_csv("C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/previous_data/全店3.csv",encoding='cp932')#過去実績翌週
previous_week3_sales_values = pd.read_csv('C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/sales_values/売上実績3.csv',encoding='cp932')#来週売上集計
tenpo = [
    ["01001008 FUN柏","柏"],
    ["01001009 FUN千葉C-one","千葉"],
    ["01001028 FUNスマーク伊勢崎","伊勢崎"],
    # ["01001032 FUNララガーデン長町","長町"],
    # ["01001033 FUNららぽーとTOKYO-BAY","船橋"],
    ["01001034 FUNららぽーと富士見","富士見"],
    ["01001036 FUNイオンレイクタウン","レイク"],
    ["01001038 FUNららぽーと海老名","海老名"],
    ["01001039 FUNイオンモールむさし村山","むさし"],
    ["01001040 FUNららぽーと湘南平塚","平塚"],
    ["01001041 FUNイオンモール名取","名取"],
    ["01001042 FUNイオンモール大高","大高"],
    ["01001043 FUNららぽーと愛知東郷","東郷町"],
    ["01001044 FUNイオンモール太田","太田"],
    ["01001045 FUNイオンモール水戸内原","水戸"],
    ["01001046 FUNららぽーとEXPOCITY","EXPO"],
    ["01001047 FUNラゾーナ川崎プラザ","川崎"],
    ["01001048 FUNららぽーと新三郷","新三郷"],
    ["01001049 FUNイオンモール幕張新都心","幕張"],
    ["01001050 FUNイオンモール各務原","各務原"],
    ["01001051 FUNららぽーと堺","堺"],
    
]

item_cd_list = {
  "OP":"01",                                                                                                                                       
  "CD":"02",
  "JK":"03",
  "KT":"04",
  "CS":"05",
  "CT":"06",
  "BL":"07",
  "SK":"08",
  "PT":"09",
  "TR":"10",
  "INN":"11",
  "SETUP":"12",
  "ACC":"13",
  "SH":"15",
}

lowest_line = 0.9

comment_1 = "全店実績の" + "{: .1f}".format(lowest_line) + "%減が許容最低ラインになります。"

print(comment_1)
path_1 = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/data_folder/'
path_2 = 'C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/'#保存先
#range_ = 0

output_faile = ["C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/週間分析.xlsx","週次","商品実績","Sheet2","柏"]#パス/Sheet Nam

wb = pyxl.load_workbook(output_faile[0])
ws_writer_1 = wb[output_faile[1]]


cell_1 = ws_writer_1.cell(2,5).value
cell_2 = ws_writer_1.cell(3,5).value
cell_3 = ws_writer_1.cell(4,5).value
cell_4 = ws_writer_1.cell(5,5).value
cell_5 = ws_writer_1.cell(6,5).value
cell_6 = ws_writer_1.cell(7,5).value
cell_7 = ws_writer_1.cell(8,5).value
cell_8 = ws_writer_1.cell(9,5).value
cell_9 = ws_writer_1.cell(10,5).value
cell_10 = ws_writer_1.cell(11,5).value
cell_11 = ws_writer_1.cell(12,5).value
cell_12 = ws_writer_1.cell(13,5).value
cell_13 = ws_writer_1.cell(14,5).value


cell_list = [
  
  cell_1,
  cell_2,
  cell_3,
  cell_4,
  cell_5,
  cell_6,
  cell_7,
  cell_8,
  cell_9,
  cell_10,
  cell_11,
  cell_12,
  cell_13, 
  
]





try :
  all_values = sum(cell_list)
except :
  cell_1 = 0
  cell_2 = 0
  cell_3 = 0
  cell_4 = 0
  cell_5 = 0
  cell_6 = 0
  cell_7 = 0
  cell_8 = 0
  cell_9 = 0
  cell_10 = 0
  cell_11 = 0
  cell_12 = 0
  cell_13 = 0


  cell_list = [
    
    cell_1,
    cell_2,
    cell_3,
    cell_4,
    cell_5,
    cell_6,
    cell_7,
    cell_8,
    cell_9,
    cell_10,
    cell_11,
    cell_12,
    cell_13, 
    
  ]
  
  all_values = sum(cell_list)
  

lowest_line_all_v = all_values * lowest_line

difference = all_values - lowest_line_all_v #差分

print(all_values)
print(lowest_line_all_v)

item_ratio = []

for n in range(0,13):
  #ratio = "{: .1f}".format(cell_list[n]/all_values * 100)
  try :
    ratio = cell_list[n]/all_values
    
  except ZeroDivisionError:
    ratio = 0
      
  diff_ratio = difference * ratio
  item_ratio.append(diff_ratio)
  
  diff_ratio_1 = cell_list[n] - diff_ratio
  
  ws_writer_1["F" + str(2 + n)].value = diff_ratio_1#追加

for cell_no in range(0,13):
  data_list = []
  avg_list = []
  under_list = []#標準区間以下
  over_list = []#標準区間以上

  for sheet_n in tenpo:
    print(sheet_n[1])
    ws = wb[str(sheet_n[1])]

    category_names = ws["B" + str(17 + cell_no)].value#　8 ⇒　17　に変更
    print(category_names)
    #cell_sales = ws["C" + str(8 + cell_no)].value
    #cell_ratio = ws["D" + str(8 + cell_no)].value
    cell_v = ws["E" + str(17 + cell_no)].value#　8 ⇒　17　に変更
    
    data_list.append(cell_v)
  #★  
  std_1 = np.std(data_list)
  avg1 = np.average(data_list)

  for data_n in data_list:
    pin1 = avg1 - std_1
    pin2 = avg1 + std_1
    
    if (pin1 < data_n) & (pin2 > data_n):  
      avg_list.append(data_n)
      
    elif pin1 > data_n :
      
      under_list.append(data_n)
      
    elif pin2 < data_n :
      over_list.append(data_n)  
      
    
  print("標準偏差",std_1)  
  #print(avg_list)
  print("区間平均",np.average(avg_list))
  
  #ws_writer_1["F" + str(2 + cell_no)].value = pin1#標準区間最低値
  #標準区間最低値を下回る数値のアイテムを改善アイテムとして提案
  
  
  ws_writer_1["G" + str(2 + cell_no)].value = np.average(avg_list)#標準期間平均(トリム平均)
  
  ws_writer_1["H" + str(2 + cell_no)].value = pin2#標準区間最大値
  
  
  print( "【標準区間】"  + "\n" + str(pin1) + " 〜 " + str(pin2))
  print("標準区間以下 ⇒ " + str(len(under_list)) + "\n" + "標準区間内⇒ " + str(len(avg_list)) + "\n" + "標準区間以上 ⇒ " + str(len(over_list)))


wb.save(output_faile[0])

wb_1 = pyxl.load_workbook(output_faile[0],data_only=True)

#◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆

#wb_1.save(path_2 + "テスト週間分析.xlsx")
wb_1.save(path_2 + create_filename)

wb_1 = pyxl.load_workbook(path_2 + create_filename)

#◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆

for sheet_names in tenpo:
  print(sheet_names)
  #range_ = 0
  
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  #全店実績のランキングデータを作成　★★★ START ★★★ 
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  all_rank_data = pd.read_csv(str(path_1) + '全店.csv',encoding="SHIFT-JIS")
  df_all_rank_data = pd.DataFrame(all_rank_data)
  
  item_cd = pd.DataFrame(df_all_rank_data["商品コード"].astype('str').str.zfill(10).values,columns=["商品CD"])
  item_name = pd.DataFrame(df_all_rank_data["商品名"].values,columns=["商品名"])
  category_cd = pd.DataFrame(df_all_rank_data["商品コード"].astype('str').str.zfill(10).str[2:4].values,columns=["アイテムCD"])
  quantity = pd.DataFrame(df_all_rank_data['合計数量'].values,columns=["数量"])
  amount = pd.DataFrame(df_all_rank_data['合計金額'].values,columns=["金額"])
  #shop_name = pd.DataFrame([shop_i[2]],columns=["店舗"])

  df_all_rank_data_values = pd.concat([item_cd,item_name,category_cd,quantity,amount],axis=1)

  filter1_df_all_rank_data_values = df_all_rank_data_values[df_all_rank_data_values["アイテムCD"] != "98" ]

  filter2_df_all_rank_data_values = filter1_df_all_rank_data_values[(filter1_df_all_rank_data_values["商品名"] != "ｷﾚｲﾏｽｸ") & (filter1_df_all_rank_data_values["商品名"] != "ｻﾝﾌﾟﾙ") ]
  
  filter2_df_all_rank_data_values2 = pd.DataFrame(filter2_df_all_rank_data_values)

  
  
  
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  #全店実績のランキングデータを作成　★★★ END ★★★ 
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  
  #＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞＞
  
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  #店別実績のランキングデータを作成　★★★ START ★★★ 
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  shop_rank_data = pd.read_csv(str(path_1) + str(sheet_names[1]) + '.csv',encoding="SHIFT-JIS")
  df_shop_rank_data = pd.DataFrame(shop_rank_data)
  
  item_cd = pd.DataFrame(df_shop_rank_data["商品コード"].astype('str').str.zfill(10).values,columns=["商品CD"])
  item_name = pd.DataFrame(df_shop_rank_data["商品名"].values,columns=["商品名"])
  category_cd = pd.DataFrame(df_shop_rank_data["商品コード"].astype('str').str.zfill(10).str[2:4].values,columns=["アイテムCD"])
  quantity = pd.DataFrame(df_shop_rank_data['合計数量'].values,columns=["数量"])
  amount = pd.DataFrame(df_shop_rank_data['合計金額'].values,columns=["金額"])
  #shop_name = pd.DataFrame([shop_i[2]],columns=["店舗"])

  df_shop_rank_data_values = pd.concat([item_cd,item_name,category_cd,quantity,amount],axis=1)

  filter1_df_shop_rank_data_values = df_shop_rank_data_values[df_shop_rank_data_values["アイテムCD"] != "98" ]

  filter2_df_shop_rank_data_values = filter1_df_shop_rank_data_values[(filter1_df_shop_rank_data_values["商品名"] != "ｷﾚｲﾏｽｸ") & (filter1_df_shop_rank_data_values["商品名"] != "ｻﾝﾌﾟﾙ") ]
  
  filter2_df_shop_rank_data_values2 = pd.DataFrame(filter2_df_shop_rank_data_values)
  
  
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  #店別実績のランキングデータを作成　★★★ END ★★★ 
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  
  #output_faile = ["C:/Users/fun-f/Desktop/analysis/週間分析.xlsx","週次","商品実績","Sheet2","柏"]#パス/Sheet Name

  sheet_data = pd.read_excel(output_faile[0],sheet_name=sheet_names[1])


  #wb_1 = pyxl.load_workbook(output_faile[0],data_only=True)
  #wb_1.active
  
  #◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆
  
  #wb_1.save(path_2 + "テスト週間分析.xlsx")
  
  #wb_1 = pyxl.load_workbook(path_2 + "テスト週間分析.xlsx")
  
  #◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆◆
  ws_1 = wb_1[str(sheet_names[1])]#店別シート
  ws_2 = wb_1[str(output_faile[1])]#全店実績シート

  weak_point = []
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  #店舗のウィークポイントを判定　★★★ START ★★★　
  #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
  #全店と店別の実績比較をして問題のあるカテゴリーを抽出する

  for gain in range(0,12):
    ws_3_cellelements = ws_1["B" + str(17 + gain)].value
    
    ws_1_cellelements = ws_1["F" + str(17 + gain)].value
    ws_2_cellelements = ws_2["F" + str(2 + gain)].value
    out_element1 = ws_1["G" + str(17 + gain)].value
    #
      #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>

        
    if out_element1 > 10 :#20000
    
    #ws_1_cellelements = ws_1["D" + str(7 + gain)].value
    #ws_2_cellelements = ws_2["D" + str(2 + gain)].value

      #インデント
      try :
        elements3 = (ws_1_cellelements - ws_2_cellelements) * 100
      except :
        elements3 = 0
      
      elements4 = pd.DataFrame({
        "アイテム名": [ws_3_cellelements],
        "差異":[elements3]
      })
      
      if elements3 < 0:
    
        weak_point.append(elements4)
      
  try:
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    #　　　　　　■■■■■■■■■■■■■■■■■■■■■■■■　WeakPoint ■■■■■■■■■■■■■■■■■■■■■■■■
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    
    concat_weak_point = pd.concat(weak_point).sort_values("差異",ascending=True).head(5) 
    
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
    
    #>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>>
  
    
 # except ValueError:
    
    
    print(concat_weak_point)
    
    
    #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
    #店別ウィークポイントを出力　★★★ END ★★★
    #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
    header_list_non = [
      ["B","C","D","E"],
      ["F","G","H","I"],#アイテムCD,[比較値【％】/商品名],[全店実績/商品名],[自店実績/SET販売数]
      ["J","K","L","M"],
      ["N","O","P","Q"],
      ["R","S","T","U"],
    ]
    
    header_list = [
      ["C","D","H","I"],
      ["K","L","P","Q"],#アイテムCD,[比較値【％】/商品名],[全店実績/商品名],[自店実績/SET販売数]
      ["S","T","Z","AB"],
      ["AF","AH","AN","AP"],
      ["AT","AV","BB","BD"],
    ]
    
    header_cont = 0#初期値
    for ac in concat_weak_point["アイテム名"]:
      weak_elements1 = item_cd_list[ac]
      #以下　アイテム別ランキング

      customer_data = pd.read_csv('C:/Users/古内翔平/OneDrive - 株式会社　ＴＲＩＮＩＴＹ　/業務会議/4⃣販売部/古内/analysis/data_folder/全店顧客データ.csv',encoding='cp932')#今週実績
      df_customer_data = pd.DataFrame(customer_data)

      df_week1 = pd.DataFrame(week1)#前週実績
      df_week1_sales_values = pd.DataFrame(week1_sales_values)

      #--------------------------------------------------------------------------------------------------------------------
      #顧客販売データを処理
      order_n = pd.DataFrame(df_customer_data["伝票番号"],columns=["伝票番号"])
      item_cd = pd.DataFrame(df_customer_data["商品コード"].astype('str').str.zfill(10).str[:10].values,columns=["商品CD"])
      item_name = pd.DataFrame(df_customer_data["商品名"],columns=["商品名"])
      category_cd = pd.DataFrame(df_customer_data["商品コード"].astype('str').str.zfill(10).str[2:4].values,columns=["アイテムCD"])
      quantity = pd.DataFrame(df_customer_data["数量"].values,columns=["数量"])#伝票明細数量
      amount = pd.DataFrame(df_customer_data["小計金額"].values,columns=["金額"])#伝票明細小計金額

      set_data = pd.concat([order_n,item_cd,item_name,category_cd,quantity,amount],axis=1)
        
      #filter_data = set_data[set_data["金額"] >= 100]

      filter_1 = set_data[set_data["アイテムCD"] != "98"] #ショッパー除外

      filter_2 = filter_1[filter_1["アイテムCD"] != "14"] #サンプル除外

      filter_data = filter_2[filter_2["商品名"] != "ｷﾚｲﾏｽｸ"] #マスク除外

      #--------------------------------------------------------------------------------------------------------------------

      #print(df_week1)

      noc = sum(df_week1_sales_values["売上客数"].values)#売上客数
      
      #全店実績

      item_cd = pd.DataFrame(df_week1["商品コード"].astype('str').str.zfill(10).values,columns=["商品CD"])
      item_name = pd.DataFrame(df_week1["商品名"].values,columns=["商品名"])
      category_cd = pd.DataFrame(df_week1["商品コード"].astype('str').str.zfill(10).str[2:4].values,columns=["アイテムCD"])
      quantity = pd.DataFrame(df_week1['合計数量'].values,columns=["数量"])
      amount = pd.DataFrame(df_week1['合計金額'].values,columns=["金額"])


      df_week1_values = pd.concat([item_cd,item_name,category_cd,quantity,amount],axis=1)

      filter1_df_week1_values = df_week1_values[df_week1_values["アイテムCD"] != "98" ]

      filter2_df_week1_values = filter1_df_week1_values[(filter1_df_week1_values["商品名"] != "ｷﾚｲﾏｽｸ") & (filter1_df_week1_values["商品名"] != "ｻﾝﾌﾟﾙ") ]


      all_amount = sum(filter2_df_week1_values["金額"].values)
      
      #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
      #アイテム別全店実績ランキング５位
      
      rank_data1 = filter2_df_week1_values[filter2_df_week1_values["アイテムCD"] == weak_elements1].sort_values("金額",ascending=False).head(5)#★★★
      
      list_count = len(rank_data1["商品名"].values)
      print(list_count)
      #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv
    
      
      
      data_list = []
      del_list = []
      
      start_values = 0
    
      writer_ws = wb_1[str(sheet_names[1])]#出力先シート名を設定
      #for writer1 in concat_weak_point.values:
        #print(writer1)
        
      col_1  = "R"
      col_2  = "S"
        
      try :  
        writer_ws[col_1 + str(5)].value = concat_weak_point.values[0][0]
        
      except IndexError:
        writer_ws[col_1 + str(5)].value = ""
        
      try :  
        writer_ws[col_1 + str(6)].value = concat_weak_point.values[1][0]
        
      except IndexError:
        writer_ws[col_1 + str(6)].value = ""  
        
      try :  
        writer_ws[col_1 + str(7)].value = concat_weak_point.values[2][0]
        
      except IndexError:
        writer_ws[col_1 + str(7)].value = ""
    
      try :  
        writer_ws[col_1 + str(8)].value = concat_weak_point.values[3][0]
        
      except IndexError:
        writer_ws[col_1 + str(8)].value = ""
        
        
      try :  
        writer_ws[col_1 + str(9)].value = concat_weak_point.values[4][0]
        
      except IndexError:
        writer_ws[col_1 + str(9)].value = ""  
        
    # try :  
        #writer_ws[col_1 + str(13)].value = concat_weak_point.values[5][0]
        
      #except IndexError:
        #writer_ws[col_1 + str(13)].value = ""  
        
      #vvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvvv END  
      try:
        writer_ws[col_2 + str(5)].value = concat_weak_point.values[0][1]
        
      except IndexError:
        writer_ws[col_2 + str(5)].value = ""
        
      try:
        writer_ws[col_2 + str(6)].value = concat_weak_point.values[1][1]
        
      except IndexError:
        writer_ws[col_2 + str(6)].value = ""
        
      try:
        writer_ws[col_2 + str(7)].value = concat_weak_point.values[2][1]
        
      except IndexError:
        writer_ws[col_2 + str(7)].value = ""    
        
      try:
        writer_ws[col_2 + str(8)].value = concat_weak_point.values[3][1]
        
      except IndexError:
        writer_ws[col_2 + str(8)].value = ""  
        
      try:
        writer_ws[col_2 + str(9)].value = concat_weak_point.values[4][1]
        
      except IndexError:
        writer_ws[col_2 + str(9)].value = ""  
        
      #try:
        #writer_ws[col_2 + str(13)].value = concat_weak_point.values[5][1]
        
      #except IndexError:
        #writer_ws[col_2 + str(31)].value = ""  
    
        start_values += 1
      
      
      #wb_1.save(output_faile[0])#◆◆◆
      #wb_1.save(path_2 + "テスト週間分析.xlsx")
      
      ####################################################################################################################################
      #for rank_n in rank_data1["商品名"].values:
      writer_ws = wb_1[str(sheet_names[1])]#出力先シート名を設定
      
      try:
        rank_1 = rank_data1["商品名"].values[0]
        print(rank_1)
        
        shop_item_values1 = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == rank_1]#アイt無名が一致するアイテムを作成
        ranking_1_shop_values = shop_item_values1["数量"].values#★ランキング実績上位５品番の店別実績を出力
        
        macth_data1 = filter_data[filter_data["商品名"] == rank_1]

        order_no1 = np.unique(macth_data1["伝票番号"].values)
        
        quantity_data1 = len(order_no1)#販売点数
      
        
        writer_ws = wb_1[str(sheet_names[1])]#出力先シート名を設定
        writer_ws[header_list[header_cont][0] + str(84)].value = rank_1
        writer_ws[header_list[header_cont][2] + str(84)].value = quantity_data1
        
        try:
          writer_ws[header_list[header_cont][3] + str(84)].value = ranking_1_shop_values[0]
        except IndexError:
          writer_ws[header_list[header_cont][3] + str(84)].value = 0
          

        #-------------------------------------------------------------------
        
        list1_1 = []
        for order_x in order_no1:
      
          datas = filter_data[filter_data["伝票番号"] == order_x]  
          
          if len(datas) > 1 :
            
            for datas_contents in datas.values:
              
              #shop_item_values = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == datas_contents[2]]
              #print(shop_item_values["数量"].values)
              
              df_datas_contents = pd.DataFrame({
                "伝票番号":[datas_contents[0]],
                "商品CD":datas_contents[1],
                "商品名":datas_contents[2],
                "アイテムCD":datas_contents[3],
                "数量":datas_contents[4],
                "金額":datas_contents[5],
                
                })
              
              list1_1.append(df_datas_contents)
              #print(datas_contents)
        list1_1_concat_1 = pd.concat(list1_1)  
        list1_1_concat_2 = list1_1_concat_1[list1_1_concat_1["商品名"] != rank_1]
        
        counts = list1_1_concat_2["商品名"].value_counts().head(5)

        list1_2 = []
        for a,b in zip(counts.index,counts):
          
          mac_element = list1_1_concat_2[list1_1_concat_2["商品名"] == a ]
          
          
          #アイテムCDを取得
          mac_element_2 = mac_element["アイテムCD"].values[0]
          
          #商品CDを取得
          mac_element_3 = mac_element["商品CD"].values[0]

          ranking_datas = pd.DataFrame({

                "商品CD":[mac_element_3],
                "商品名":[a],
                "アイテムCD":[mac_element_2],
                "数量":[b],
                #"金額":datas_contents[5]
                })
          
          list1_2.append(ranking_datas)
          
        list1_2_concat = pd.concat(list1_2)#SET実績ベスト５を作成
        
        start_row = 86
        writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list1_2_concat.values[0][1]
        writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list1_2_concat.values[1][1]
        writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list1_2_concat.values[2][1]
        writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list1_2_concat.values[3][1]
        writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list1_2_concat.values[4][1]
        
        writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list1_2_concat.values[0][3]
        writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list1_2_concat.values[1][3]
        writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list1_2_concat.values[2][3]
        writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list1_2_concat.values[3][3]
        writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list1_2_concat.values[4][3]

        #-------------------------------------------------------------------
        
      except IndexError:
        ranking_1_shop_values = ""#★ランキング実績上位５品番の店別実績を出力
        
        if list_count < 1 :
          
          start_row = 86
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
          
        else :  
        
        
          start_row = 86
          
          try:
           writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list1_2_concat.values[0][1]
          except IndexError:
           writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = "" 
             
          try :     
            writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list1_2_concat.values[1][1]    
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
            
          try :  
            writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list1_2_concat.values[2][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
            
          try :  
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list1_2_concat.values[3][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
            
            
          try :
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list1_2_concat.values[4][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
            
          try :
            writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list1_2_concat.values[0][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
              
          try :  
            writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list1_2_concat.values[1][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
            
          try :  
            writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list1_2_concat.values[2][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
            
          try :  
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list1_2_concat.values[3][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          
          try :  
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list1_2_concat.values[4][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
            
          
    
        
    
      
      try:
      
        rank_2 = rank_data1["商品名"].values[1]
        
        shop_item_values2 = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == rank_2]#アイt無名が一致するアイテムを作成
        ranking_2_shop_values = shop_item_values2["数量"].values#★ランキング実績上位５品番の店別実績を出力
      
        
        macth_data2 = filter_data[filter_data["商品名"] == rank_2]

        order_no2 = np.unique(macth_data2["伝票番号"].values)
        
        quantity_data2 = len(order_no2)#販売点数
        
        
        
        writer_ws = wb_1[str(sheet_names[1])]#出力先シート名を設定
        
        writer_ws[header_list[header_cont][0] + str(91)].value = rank_2
        writer_ws[header_list[header_cont][2] + str(91)].value = quantity_data2
        try :
          writer_ws[header_list[header_cont][3] + str(91)].value = ranking_2_shop_values[0]
        except IndexError:
          
          writer_ws[header_list[header_cont][3] + str(91)].value = 0
          
        
        #-------------------------------------------------------------------
        
        list2_1 = []
        for order_x in order_no2:
      
          datas = filter_data[filter_data["伝票番号"] == order_x]  
          
          if len(datas) > 1 :
            
            for datas_contents in datas.values:
              
              #shop_item_values = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == datas_contents[2]]
              #print(shop_item_values["数量"].values)
              
              df_datas_contents = pd.DataFrame({
                "伝票番号":[datas_contents[0]],
                "商品CD":datas_contents[1],
                "商品名":datas_contents[2],
                "アイテムCD":datas_contents[3],
                "数量":datas_contents[4],
                "金額":datas_contents[5],
                
                })
              
              list2_1.append(df_datas_contents)
              #print(datas_contents)
        list2_1_concat_1 = pd.concat(list2_1)  
        list2_1_concat_2 = list2_1_concat_1[list2_1_concat_1["商品名"] != rank_2]
        
        counts = list2_1_concat_2["商品名"].value_counts().head(5)

        list2_2 = []
        for a,b in zip(counts.index,counts):
          
          mac_element = list2_1_concat_2[list2_1_concat_2["商品名"] == a ]
          
          
          #アイテムCDを取得
          mac_element_2 = mac_element["アイテムCD"].values[0]
          
          #商品CDを取得
          mac_element_3 = mac_element["商品CD"].values[0]

          ranking_datas = pd.DataFrame({

                "商品CD":[mac_element_3],
                "商品名":[a],
                "アイテムCD":[mac_element_2],
                "数量":[b],
                #"金額":datas_contents[5]
                })
          
          list2_2.append(ranking_datas)
          
        list2_2_concat = pd.concat(list2_2)#SET実績ベスト５を作成
        
        try:
          
          
          
          start_row = 93
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list2_2_concat.values[0][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list2_2_concat.values[1][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list2_2_concat.values[2][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list2_2_concat.values[3][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list2_2_concat.values[4][1]
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list2_2_concat.values[0][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list2_2_concat.values[1][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list2_2_concat.values[2][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list2_2_concat.values[3][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list2_2_concat.values[4][3]
          
        except IndexError:
          
          start_row = 93
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
          
  
        #-------------------------------------------------------------------
        
      except IndexError:
        ranking_2_shop_values = ""#★ランキング実績上位５品番の店別実績を出力
        
        if list_count < 2 :
          start_row = 93
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
          
        
        else :
          start_row = 93
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list2_2_concat.values[0][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list2_2_concat.values[1][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list2_2_concat.values[2][1]
          try :
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list2_2_concat.values[3][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          
          try :  
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list2_2_concat.values[4][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list2_2_concat.values[0][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list2_2_concat.values[1][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list2_2_concat.values[2][3]
          try :
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list2_2_concat.values[3][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value =""
            
          try :  
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list2_2_concat.values[4][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
            
        
      
      try:
        rank_3 = rank_data1["商品名"].values[2]
        
      
        shop_item_values3 = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == rank_3]#アイt無名が一致するアイテムを作成
        ranking_3_shop_values = shop_item_values3["数量"].values#★ランキング実績上位５品番の店別実績を出力
        
      
        
        macth_data3 = filter_data[filter_data["商品名"] == rank_3]

        order_no3 = np.unique(macth_data3["伝票番号"].values)
        
        quantity_data3 = len(order_no3)#販売点数
        
        writer_ws = wb_1[str(sheet_names[1])]#出力先シート名を設定
        writer_ws[header_list[header_cont][0] + str(98)].value = rank_3
        writer_ws[header_list[header_cont][2] + str(98)].value = quantity_data3
        
        try :
          writer_ws[header_list[header_cont][3] + str(98)].value = ranking_3_shop_values[0]
          
        except IndexError:
          writer_ws[header_list[header_cont][3] + str(98)].value = 0
      
        
        #-------------------------------------------------------------------
        
        list3_1 = []
        for order_x in order_no3:
      
          datas = filter_data[filter_data["伝票番号"] == order_x]  
          
          if len(datas) > 1 :
            
            for datas_contents in datas.values:
              
              #shop_item_values = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == datas_contents[2]]
              #print(shop_item_values["数量"].values)
              
              df_datas_contents = pd.DataFrame({
                "伝票番号":[datas_contents[0]],
                "商品CD":datas_contents[1],
                "商品名":datas_contents[2],
                "アイテムCD":datas_contents[3],
                "数量":datas_contents[4],
                "金額":datas_contents[5],
                
                })
              
              list3_1.append(df_datas_contents)
              #print(datas_contents)
        list3_1_concat_1 = pd.concat(list3_1)  
        list3_1_concat_2 = list3_1_concat_1[list3_1_concat_1["商品名"] != rank_3]
        
        counts = list3_1_concat_2["商品名"].value_counts().head(5)

        list3_2 = []
        for a,b in zip(counts.index,counts):
          
          mac_element = list3_1_concat_2[list3_1_concat_2["商品名"] == a ]
          
          
          #アイテムCDを取得
          mac_element_2 = mac_element["アイテムCD"].values[0]
          
          #商品CDを取得
          mac_element_3 = mac_element["商品CD"].values[0]

          ranking_datas = pd.DataFrame({

                "商品CD":[mac_element_3],
                "商品名":[a],
                "アイテムCD":[mac_element_2],
                "数量":[b],
                #"金額":datas_contents[5]
                })
          
          list3_2.append(ranking_datas)
          
        list3_2_concat = pd.concat(list3_2)#SET実績ベスト５を作成
        
        print(list3_2_concat)
        
        
        start_row = 100
        try :
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list3_2_concat.values[0][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list3_2_concat.values[1][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list3_2_concat.values[2][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list3_2_concat.values[3][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list3_2_concat.values[4][1]
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list3_2_concat.values[0][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list3_2_concat.values[1][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list3_2_concat.values[2][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list3_2_concat.values[3][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list3_2_concat.values[4][3]
          
        except IndexError:
          
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
            
    
        
        
        
        #-------------------------------------------------------------------
        
      except IndexError:
        ranking_3_shop_values = ""#★ランキング実績上位５品番の店別実績を出力
        
        if list_count < 3 :
          start_row = 100
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
          
        else:
        
          start_row = 100
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list3_2_concat.values[0][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list3_2_concat.values[1][1]
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list3_2_concat.values[2][1]
          try :
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list3_2_concat.values[3][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
            
          try :  
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list3_2_concat.values[4][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
            
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list3_2_concat.values[0][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list3_2_concat.values[1][3]
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list3_2_concat.values[2][3]
          try :
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list3_2_concat.values[3][3]
          except IndexError :
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          
          try :  
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list3_2_concat.values[4][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
      

      
      
      try:
        
        rank_4 = rank_data1["商品名"].values[3]
      
        shop_item_values4 = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == rank_4]#アイt無名が一致するアイテムを作成
        ranking_4_shop_values = shop_item_values4["数量"].values#★ランキング実績上位５品番の店別実績を出力
      
        
        macth_data4 = filter_data[filter_data["商品名"] == rank_4]

        order_no4 = np.unique(macth_data4["伝票番号"].values)
        
        quantity_data4 = len(order_no4)#販売点数
      
        
        writer_ws = wb_1[str(sheet_names[1])]#出力先シート名を設定
        writer_ws[header_list[header_cont][0] + str(105)].value = rank_4
        writer_ws[header_list[header_cont][2] + str(105)].value = quantity_data4
        
        try:
          writer_ws[header_list[header_cont][3] + str(105)].value = ranking_4_shop_values[0]
          
        except IndexError:
          
          writer_ws[header_list[header_cont][3] + str(105)].value = 0
          
        
        #-------------------------------------------------------------------
        
        list4_1 = []
        for order_x in order_no4:
      
          datas = filter_data[filter_data["伝票番号"] == order_x]  
          
          if len(datas) > 1 :
            
            for datas_contents in datas.values:
              
              #shop_item_values = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == datas_contents[2]]
              #print(shop_item_values["数量"].values)
              
              df_datas_contents = pd.DataFrame({
                "伝票番号":[datas_contents[0]],
                "商品CD":datas_contents[1],
                "商品名":datas_contents[2],
                "アイテムCD":datas_contents[3],
                "数量":datas_contents[4],
                "金額":datas_contents[5],
                
                })
              
              list4_1.append(df_datas_contents)
              #print(datas_contents)
        list4_1_concat_1 = pd.concat(list4_1)  
        list4_1_concat_2 = list4_1_concat_1[list4_1_concat_1["商品名"] != rank_4]
        
        counts = list4_1_concat_2["商品名"].value_counts().head(5)

        list4_2 = []
        for a,b in zip(counts.index,counts):
          
          mac_element = list4_1_concat_2[list4_1_concat_2["商品名"] == a ]
          
          
          #アイテムCDを取得
          mac_element_2 = mac_element["アイテムCD"].values[0]
          
          #商品CDを取得
          mac_element_3 = mac_element["商品CD"].values[0]

          ranking_datas = pd.DataFrame({

                "商品CD":[mac_element_3],
                "商品名":[a],
                "アイテムCD":[mac_element_2],
                "数量":[b],
                #"金額":datas_contents[5]
                })
          
          list4_2.append(ranking_datas)
          
        list4_2_concat = pd.concat(list4_2)#SET実績ベスト５を作成
        
        start_row = 107
        try:
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list4_2_concat.values[0][1]
        except IndexError:
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
        
        try:    
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list4_2_concat.values[1][1]
        except IndexError:
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
            
        try:  
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list4_2_concat.values[2][1]
        except IndexError:
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
        
        try:    
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list4_2_concat.values[3][1]
          
        except IndexError:
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
            
        try:    
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list4_2_concat.values[4][1]
        except IndexError:
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
            
        try:
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list4_2_concat.values[0][3]
          
        except IndexError:
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
        
        try:    
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list4_2_concat.values[1][3]
        except IndexError:
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
        try:    
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list4_2_concat.values[2][3]
        except IndexError:
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value =""
            
        try:    
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list4_2_concat.values[3][3]
        except IndexError:
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
        
        try:  
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list4_2_concat.values[4][3]
        except IndexError:
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
            
    
        
      
        
        #-------------------------------------------------------------------
        
      except IndexError:
        ranking_4_shop_values = ""#★ランキング実績上位５品番の店別実績を出力
        
        if list_count < 4 :
          
          start_row = 107
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
        
        else :  
      
        
          start_row = 107
          try:
            writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list4_2_concat.values[0][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          
          try:    
            writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list4_2_concat.values[1][1]
            
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
        
          try:  
            writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list4_2_concat.values[2][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
              
            
          try :
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list4_2_concat.values[3][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          try:  
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list4_2_concat.values[4][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
            
          try:  
            
            writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list4_2_concat.values[0][3]
            
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          try:    
            writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list4_2_concat.values[1][3]
            
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
          
          try:    
            writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list4_2_concat.values[2][3]
            
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
              
          try :
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list4_2_concat.values[3][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          try :  
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list4_2_concat.values[4][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
        
        
      
      try:
      
        rank_5 = rank_data1["商品名"].values[4]
      
        shop_item_values5 = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == rank_5]#アイt無名が一致するアイテムを作成
        ranking_5_shop_values = shop_item_values5["数量"].values#★ランキング実績上位５品番の店別実績を出力
      
        
        macth_data5 = filter_data[filter_data["商品名"] == rank_5]

        order_no5 = np.unique(macth_data5["伝票番号"].values)
        
        quantity_data5 = len(order_no5)#販売点数
        

        
        writer_ws = wb_1[str(sheet_names[1])]#出力先シート名を設定
        writer_ws[header_list[header_cont][0] + str(112)].value = rank_5
        writer_ws[header_list[header_cont][2] + str(112)].value = quantity_data5
        try:
          writer_ws[header_list[header_cont][3] + str(112)].value = ranking_5_shop_values[0]
          
        except IndexError:
          
          writer_ws[header_list[header_cont][3] + str(112)].value = 0
          
        
        #-------------------------------------------------------------------
        
        list5_1 = []
        for order_x in order_no5:
      
          datas = filter_data[filter_data["伝票番号"] == order_x]  
          
          if len(datas) > 1 :
            
            for datas_contents in datas.values:
              
              #shop_item_values = filter2_df_shop_rank_data_values2[filter2_df_shop_rank_data_values2["商品名"] == datas_contents[2]]
              #print(shop_item_values["数量"].values)
              
              df_datas_contents = pd.DataFrame({
                "伝票番号":[datas_contents[0]],
                "商品CD":datas_contents[1],
                "商品名":datas_contents[2],
                "アイテムCD":datas_contents[3],
                "数量":datas_contents[4],
                "金額":datas_contents[5],
                
                })
              
              list5_1.append(df_datas_contents)
              #print(datas_contents)
        list5_1_concat_1 = pd.concat(list5_1)  
        list5_1_concat_2 = list5_1_concat_1[list5_1_concat_1["商品名"] != rank_5]
        
        counts = list5_1_concat_2["商品名"].value_counts().head(5)

        list5_2 = []
        for a,b in zip(counts.index,counts):
          
          mac_element = list5_1_concat_2[list5_1_concat_2["商品名"] == a ]
          
          
          #アイテムCDを取得
          mac_element_2 = mac_element["アイテムCD"].values[0]
          
          #商品CDを取得
          mac_element_3 = mac_element["商品CD"].values[0]

          ranking_datas = pd.DataFrame({

                "商品CD":[mac_element_3],
                "商品名":[a],
                "アイテムCD":[mac_element_2],
                "数量":[b],
                #"金額":datas_contents[5]
                })
          
          list5_2.append(ranking_datas)
          
        list5_2_concat = pd.concat(list5_2)#SET実績ベスト５を作成
        
        start_row = 114
        try :
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list5_2_concat.values[0][1]
          
        except IndexError:
          
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          
        try:  
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list5_2_concat.values[1][1]
        
        except IndexError:
          
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
          
        try:  
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list5_2_concat.values[2][1]
          
        except IndexError:
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
        
        
        try:  
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list5_2_concat.values[3][1]
          
        except IndexError:
          
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          
        try:  
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list5_2_concat.values[4][1]
          
        except IndexError:
          
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          
          
          
        try :
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list5_2_concat.values[0][3]
          
        except IndexError:
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          
        try:   
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list5_2_concat.values[1][3]
          
        except IndexError:
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
       
        try:  
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list5_2_concat.values[2][3]
        except IndexError:
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
            
        try:  
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list5_2_concat.values[3][3]
        except IndexError:
            
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          
          
        try:  
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list5_2_concat.values[4][3]
          
        except IndexError:
          
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
          
        #-------------------------------------------------------------------
        
      except IndexError:
        ranking_5_shop_values = ""#★ランキング実績上位５品番の店別実績を出力
        
        if  list_count < 5 :
          
          start_row = 114
          writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          
          writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
          
        else :
        
          start_row = 114
          try:
            writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = list5_2_concat.values[0][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 0)].value = ""
          try:    
            writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = list5_2_concat.values[1][1]
            
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 1)].value = ""
          
          try:    
            writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = list5_2_concat.values[2][1]
            
          except IndexError:
            
            writer_ws[header_list[header_cont][1] + str(start_row + 2)].value = ""
              
          try :
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = list5_2_concat.values[3][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 3)].value = ""
          
          try :  
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = list5_2_concat.values[4][1]
          except IndexError:
            writer_ws[header_list[header_cont][1] + str(start_row + 4)].value = ""
          try:
            writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = list5_2_concat.values[0][3]
            
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 0)].value = ""
          
          try:  
            writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = list5_2_concat.values[1][3]
            
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 1)].value = ""
          
          try:    
            writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = list5_2_concat.values[2][3]
            
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 2)].value = ""
              
          try :
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = list5_2_concat.values[3][3]    
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 3)].value = ""
          
          try :  
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = list5_2_concat.values[4][3]
          except IndexError:
            writer_ws[header_list[header_cont][3] + str(start_row + 4)].value = ""
      
      header_cont += 1
      
      #wb_1.save(path_2 + "テスト週間分析.xlsx")
      wb_1.save(os.path.join(path_2,create_filename))
          
          #range_ += 1
          
      ####################################################################################################################################
      
      
    # wb_1.save(path_2 + "テスト週間分析.xlsx")
    
    
  except ValueError:
    print("出力データなし") 

  
